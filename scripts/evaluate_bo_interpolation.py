#!/usr/bin/env python3
# =========================================================
# scripts/evaluate_bo_interpolation.py
# =========================================================
# Evaluates the BO-interpolation methods produced by
# bo_interpolation.py, writing a clean summary at THREE levels
# instead of one mixed dump:
#
#   METHOD level  (<query>/<gt_method>/bo_interpolation_results/)
#       summary.csv                      one row per BO config
#
#   QUERY level   (<query>/)
#       bo_eval_query.csv                non-aggregate: every method's configs
#       bo_eval_query.xlsx               one TAB per method (m0/m1/m2)
#       bo_eval_query_agg.csv            aggregate: mean per BO config over methods
#
#   GLOBAL level  (<gt_root>/bo_interpolation_eval/)
#       bo_eval_all.csv                  non-aggregate: every query × method × config
#       bo_eval_all.xlsx                 one TAB per query (each with all methods)
#       bo_eval_all_agg.csv              aggregate: mean per BO config over everything
#
# For every config it computes, on two splits:
#   ALL    — all pair instances
#   UNSEEN — pairs NOT sampled (true interpolation accuracy)
# plus boundary-discovery metrics (max_qerr_found, coverage_2x/5x)
# and smoothness-estimate relative errors (S_max/S_avg/S_topk).
#
# Usage:
#   python scripts/evaluate_bo_interpolation.py                         # root from config
#   python scripts/evaluate_bo_interpolation.py gt_results_sf1_10x10_s1q0
#   python scripts/evaluate_bo_interpolation.py <gt_root>/qt5           # one query
#   python scripts/evaluate_bo_interpolation.py <gt_method_dir>         # one method
# =========================================================

import os
import sys
import json
from pathlib import Path

import numpy as np
import pandas as pd
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import config_gt


SORT_BY        = "QERR_MEDIAN_UNSEEN"
TOP_K_PERCENT  = 0.10
RESULTS_SUBDIR = "bo_interpolation_results"
EVAL_SUBDIR    = "bo_interpolation_eval"

GROUP_KEYS = ["method", "representation", "kernel", "acquisition"]


# --------------------------------------------------------- metrics
def pred_qerror(y_true, y_pred):
    yt = np.maximum(np.asarray(y_true, float), 1e-9)
    yp = np.maximum(np.asarray(y_pred, float), 1e-9)
    return np.maximum(yt / yp, yp / yt)


def compute_metrics(y_true, y_pred, prefix=""):
    yt = np.asarray(y_true, float)
    yp = np.maximum(np.asarray(y_pred, float), 1e-9)
    nan = float("nan")
    keys = ["MAE", "RMSE", "R2", "MAPE", "QERR_MEAN", "QERR_MEDIAN",
            "QERR_P90", "QERR_P95", "QERR_MAX", "WITHIN_2X", "WITHIN_5X"]
    if len(yt) == 0:
        return {f"{k}{prefix}": nan for k in keys}
    pq = pred_qerror(yt, yp)
    try:
        r2 = float(r2_score(yt, yp))
    except Exception:
        r2 = nan
    return {
        f"MAE{prefix}"        : float(mean_absolute_error(yt, yp)),
        f"RMSE{prefix}"       : float(np.sqrt(mean_squared_error(yt, yp))),
        f"R2{prefix}"         : r2,
        f"MAPE{prefix}"       : float(np.mean(np.abs((yt - yp) / np.maximum(yt, 1e-9))) * 100),
        f"QERR_MEAN{prefix}"  : float(np.mean(pq)),
        f"QERR_MEDIAN{prefix}": float(np.median(pq)),
        f"QERR_P90{prefix}"   : float(np.percentile(pq, 90)),
        f"QERR_P95{prefix}"   : float(np.percentile(pq, 95)),
        f"QERR_MAX{prefix}"   : float(np.max(pq)),
        f"WITHIN_2X{prefix}"  : float(np.mean(pq <= 2.0) * 100),
        f"WITHIN_5X{prefix}"  : float(np.mean(pq <= 5.0) * 100),
    }


def _rel_err(est, true):
    if true == 0 or np.isnan(true) or np.isnan(est):
        return float("nan")
    return abs(est - true) / true * 100.0


def _metrics_row(method, df, meta):
    has_flag    = "is_sampled" in df.columns
    unseen_mask = (df["is_sampled"] == 0) if has_flag else pd.Series(True, index=df.index)
    sampled_mask = (df["is_sampled"] == 1) if has_flag else pd.Series(False, index=df.index)

    yt_all    = df["y_true"].values
    yp_all    = df["y_pred"].values
    yt_unseen = df.loc[unseen_mask, "y_true"].values
    yp_unseen = df.loc[unseen_mask, "y_pred"].values
    yt_samp   = df.loc[sampled_mask, "y_true"].values

    max_true  = float(yt_all.max()) if len(yt_all) else float("nan")
    max_found = float(yt_samp.max()) if len(yt_samp) else float("nan")
    n2 = int((yt_all > 2.0).sum()); n5 = int((yt_all > 5.0).sum())
    f2 = int((yt_samp > 2.0).sum()) if len(yt_samp) else 0
    f5 = int((yt_samp > 5.0).sum()) if len(yt_samp) else 0

    k = max(1, int(TOP_K_PERCENT * len(yt_all)))
    s_max_t = float(np.max(yt_all)); s_avg_t = float(np.mean(yt_all))
    s_top_t = float(np.mean(np.sort(yt_all)[-k:]))
    s_max_e = float(np.max(yp_all)); s_avg_e = float(np.mean(yp_all))
    s_top_e = float(np.mean(np.sort(yp_all)[-k:]))

    row = {
        "method"        : method,
        "representation": meta.get("representation", "-"),
        "kernel"        : meta.get("kernel", "-"),
        "acquisition"   : meta.get("acquisition", "-"),
    }
    row.update(compute_metrics(yt_all,    yp_all,    "_ALL"))
    row.update(compute_metrics(yt_unseen, yp_unseen, "_UNSEEN"))
    row["budget"]           = int(meta.get("budget_pairs", -1))
    row["sample_fraction"]  = float(meta.get("sample_fraction", float("nan")))
    row["num_all"]          = int(len(yt_all))
    row["num_unseen"]       = int(len(yt_unseen))
    row["dimension"]        = int(meta.get("dimension", -1))
    row["max_qerr_true"]    = max_true
    row["max_qerr_found"]   = max_found
    row["max_found_pct"]    = 100.0 * max_found / max_true if max_true else float("nan")
    row["coverage_2x_pct"]  = 100.0 * f2 / n2 if n2 else float("nan")
    row["coverage_5x_pct"]  = 100.0 * f5 / n5 if n5 else float("nan")
    row["S_max_relerr"]     = _rel_err(s_max_e, s_max_t)
    row["S_avg_relerr"]     = _rel_err(s_avg_e, s_avg_t)
    row["S_topk_relerr"]    = _rel_err(s_top_e, s_top_t)
    return row


# --------------------------------------------------------- helpers
def _aggregate(df):
    """Mean of every numeric metric per BO config, ranked by SORT_BY."""
    keys = [k for k in GROUP_KEYS if k in df.columns]
    skip = set(keys) | {"rank", "query", "gt_method"}
    metric_cols = [c for c in df.columns
                   if c not in skip and pd.api.types.is_numeric_dtype(df[c])]
    g = df.groupby(keys, as_index=False, dropna=False)
    agg = g[metric_cols].mean(numeric_only=True)
    counts = g.size().rename(columns={"size": "n_evals"})
    agg = agg.merge(counts, on=keys)
    if SORT_BY in agg.columns:
        agg = agg.sort_values(SORT_BY).reset_index(drop=True)
        agg.insert(0, "rank", agg.index + 1)
    return agg


_HDR_FILL  = PatternFill("solid", fgColor="4472C4")
_BAND_FILL = PatternFill("solid", fgColor="EDF2FB")


def _style_ws(ws, df):
    ncols = len(df.columns)
    for j in range(1, ncols + 1):
        c = ws.cell(row=1, column=j)
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i in range(2, len(df) + 2):
        if i % 2 == 1:
            for j in range(1, ncols + 1):
                ws.cell(row=i, column=j).fill = _BAND_FILL
    for j, col in enumerate(df.columns, start=1):
        sample = df.iloc[:, j - 1].astype(str).tolist()[:200]
        maxlen = max([len(str(col))] + [len(s) for s in sample]) if sample else len(str(col))
        ws.column_dimensions[get_column_letter(j)].width = max(9, min(24, maxlen + 2))
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 26


def _write_tabbed_xlsx(path, sheets):
    """sheets: ordered dict {tab_name: df}. Best-effort (skips on file lock)."""
    sheets = {k: v for k, v in sheets.items() if v is not None and len(v)}
    if not sheets:
        return False
    try:
        with pd.ExcelWriter(path, engine="openpyxl") as xw:
            used = set()
            for name, d in sheets.items():
                sheet = (str(name)[:31] or "sheet")
                base, i = sheet, 1
                while sheet in used:
                    sheet = f"{base[:28]}_{i}"; i += 1
                used.add(sheet)
                d.to_excel(xw, sheet_name=sheet, index=False)
                _style_ws(xw.sheets[sheet], d)
        return True
    except PermissionError:
        print(f"  ! {path.name} is open — skipped (close it and re-run).")
        return False


# --------------------------------------------------------- METHOD level
def evaluate_results_dir(results_dir):
    """One gt_method dir -> summary.csv (one row per BO config)."""
    results_dir = Path(results_dir)
    rows = []
    for method in sorted(os.listdir(results_dir)):
        mdir = results_dir / method
        pred = mdir / "predictions.csv"
        if not mdir.is_dir() or not pred.exists():
            continue
        df = pd.read_csv(pred)
        if "y_true" not in df.columns or "y_pred" not in df.columns:
            continue
        meta = {}
        mf = mdir / "metadata.json"
        if mf.exists():
            with open(mf) as f:
                meta = json.load(f)
        row = _metrics_row(method, df, meta)
        rows.append(row)
        with open(mdir / "metrics.json", "w") as f:
            json.dump(row, f, indent=2)

    if not rows:
        return None
    summary = pd.DataFrame(rows)
    if SORT_BY in summary.columns:
        summary = summary.sort_values(SORT_BY).reset_index(drop=True)
        summary.insert(0, "rank", summary.index + 1)
    summary.to_csv(results_dir / "summary.csv", index=False)
    return summary


# --------------------------------------------------------- QUERY level
def evaluate_query_dir(query_dir):
    """One query dir -> non-aggregate CSV, tabbed xlsx (tab per method),
    aggregate CSV. Returns the stacked (gt_method-tagged) frame or None."""
    query_dir = Path(query_dir)
    per_method = {}      # gt_method -> summary df (tab)
    stacked = []
    for gt_method in sorted(os.listdir(query_dir)):
        rdir = query_dir / gt_method / RESULTS_SUBDIR
        if not rdir.is_dir():
            continue
        s = evaluate_results_dir(rdir)
        if s is None:
            continue
        per_method[gt_method] = s
        s2 = s.copy()
        s2.insert(0, "gt_method", gt_method)
        stacked.append(s2)

    if not stacked:
        return None

    frame = pd.concat(stacked, ignore_index=True)
    frame.to_csv(query_dir / "bo_eval_query.csv", index=False)
    _aggregate(frame).to_csv(query_dir / "bo_eval_query_agg.csv", index=False)
    _write_tabbed_xlsx(query_dir / "bo_eval_query.xlsx", per_method)
    print(f"  {query_dir.name}: {len(per_method)} methods, {len(frame)} configs")
    return frame


# --------------------------------------------------------- GLOBAL level
def run_root(gt_root):
    gt_root = Path(gt_root)
    frames, tabs = [], {}
    for query_dir in sorted(p for p in gt_root.iterdir() if p.is_dir()):
        if not any(query_dir.glob(f"*/{RESULTS_SUBDIR}")):
            continue
        f = evaluate_query_dir(query_dir)
        if f is None:
            continue
        f = f.copy()
        f.insert(0, "query", query_dir.name)
        frames.append(f)
        tabs[query_dir.name] = f.drop(columns=["query"])

    if not frames:
        print(f"No {RESULTS_SUBDIR} found under {gt_root} — run bo_interpolation.py first.")
        return None

    combined = pd.concat(frames, ignore_index=True)
    out_dir  = gt_root / EVAL_SUBDIR
    out_dir.mkdir(parents=True, exist_ok=True)

    combined.to_csv(out_dir / "bo_eval_all.csv", index=False)
    _aggregate(combined).to_csv(out_dir / "bo_eval_all_agg.csv", index=False)
    _write_tabbed_xlsx(out_dir / "bo_eval_all.xlsx", tabs)

    print(f"\nGlobal:")
    print(f"  {out_dir / 'bo_eval_all.csv'}       ({len(combined)} rows, {len(tabs)} queries)")
    print(f"  {out_dir / 'bo_eval_all_agg.csv'}   (per-config aggregate)")
    print(f"  {out_dir / 'bo_eval_all.xlsx'}      (one tab per query)")
    return combined


# allow use as a processor:  run(gt_root)
def run(gt_root):
    return run_root(gt_root)


def _default_root():
    res = config_gt.get_query_resolution(config_gt.QUERIES[0], config_gt.RUN_METHODS[0])
    return config_gt.get_main_dir(res)


if __name__ == "__main__":
    print("\n================================================")
    print("BO PAIR Q-ERROR INTERPOLATION — EVALUATION")
    print("================================================")
    arg = sys.argv[1] if len(sys.argv) > 1 else None
    if arg:
        t = Path(arg)
        if t.name == RESULTS_SUBDIR:                       # a results dir (method level)
            evaluate_results_dir(t)
        elif (t / RESULTS_SUBDIR).exists():                # a gt_method dir
            evaluate_results_dir(t / RESULTS_SUBDIR)
        elif any(t.glob(f"*/{RESULTS_SUBDIR}")):           # a query dir
            evaluate_query_dir(t)
        else:                                              # a gt-root
            run_root(t)
    else:
        run_root(_default_root())
