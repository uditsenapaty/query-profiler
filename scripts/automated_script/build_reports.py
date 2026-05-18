"""
build_reports.py — Excel report + publication-quality plots.

Reads: results/<qt>/analysis_results.json, sweep_results.json, queries/<qt>.sql
Produces: results/report.xlsx, results/plots/*.png
"""

import json, os, math, argparse
from datetime import datetime

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

Q_ORDER = ["qt1","qt2","qt3","qt4","qt5","qt6","qt7","qt8",
           "qt10","qt11","qt12","qt13","qt14","qt16","qt17","qt18","qt21"]

LABELS = {"qt1":"QT1 — Pricing Summary","qt2":"QT2 — Min Cost Supplier",
    "qt3":"QT3 — Shipping Priority","qt4":"QT4 — Order Priority",
    "qt5":"QT5 — Local Supplier Volume","qt6":"QT6 — Forecasting Revenue",
    "qt7":"QT7 — Volume Shipping","qt8":"QT8 — National Market Share",
    "qt10":"QT10 — Returned Items","qt11":"QT11 — Important Stock",
    "qt12":"QT12 — Shipping Modes","qt13":"QT13 — Customer Distribution",
    "qt14":"QT14 — Promotion Effect","qt16":"QT16 — Parts/Supplier",
    "qt17":"QT17 — Small-Quantity-Order","qt18":"QT18 — Large Volume Customer",
    "qt21":"QT21 — Suppliers Kept Orders Waiting"}

PARAMS_MAP = {"qt1":["l_extendedprice"],"qt2":["p_retailprice","ps_supplycost"],
    "qt3":["o_totalprice","l_extendedprice"],"qt4":["o_totalprice","l_extendedprice"],
    "qt5":["c_acctbal","s_acctbal"],"qt6":["l_extendedprice"],
    "qt7":["o_totalprice","c_acctbal"],"qt8":["s_acctbal","l_extendedprice"],
    "qt10":["c_acctbal","l_extendedprice"],"qt11":["ps_supplycost","s_acctbal"],
    "qt12":["o_totalprice","l_extendedprice"],"qt13":["o_totalprice"],
    "qt14":["l_extendedprice","p_retailprice"],"qt16":["p_retailprice","s_acctbal"],
    "qt17":["p_retailprice","l_extendedprice"],"qt18":["l_extendedprice","c_acctbal"],
    "qt21":["s_acctbal","l_extendedprice"]}

PAL = ["#2563eb","#dc2626","#16a34a","#ea580c","#9333ea","#0891b2",
       "#d97706","#4f46e5","#059669","#e11d48","#7c3aed","#0d9488",
       "#ca8a04","#6366f1","#15803d","#be123c","#7c2d12"]

# ── Styles ──
_b = Side(style="thin", color="D1D5DB")
G = Border(bottom=_b, right=_b, left=_b, top=_b)
HF = PatternFill("solid", fgColor="1E293B")
H2F = PatternFill("solid", fgColor="334155")
HN = Font(bold=True, color="FFFFFF", name="Arial", size=9)
TF = Font(bold=True, name="Arial", size=11, color="1E293B")
SF = Font(bold=True, name="Arial", size=9, color="475569")
BF = Font(name="Arial", size=9, color="334155")
BB = Font(bold=True, name="Arial", size=9, color="1E293B")
SQ = Font(name="Consolas", size=8, color="334155")
GN = PatternFill("solid", fgColor="DCFCE7")
RD = PatternFill("solid", fgColor="FEE2E2")
YL = PatternFill("solid", fgColor="FEF3C7")
BL = PatternFill("solid", fgColor="DBEAFE")
GY = PatternFill("solid", fgColor="F1F5F9")
P1F = PatternFill("solid", fgColor="EDE9FE")
P2F = PatternFill("solid", fgColor="FEF9C3")

def _hdr(ws, r, cols, fill=HF):
    for c, h in enumerate(cols, 1):
        cl = ws.cell(row=r, column=c, value=h)
        cl.fill, cl.font, cl.border = fill, HN, G
        cl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _rw(ws, r, vals, fmts=None):
    for c, v in enumerate(vals, 1):
        cl = ws.cell(row=r, column=c, value=v)
        cl.font, cl.border = BF, G
        cl.alignment = Alignment(horizontal="center", vertical="center")
        if fmts and c-1 < len(fmts) and fmts[c-1]: cl.number_format = fmts[c-1]

def _ww(ws, widths):
    for c, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(c)].width = w

def _pv(vals):
    if not vals: return ""
    return ", ".join(f"{v:.2f}" for v in vals)


# ═══════════════════════════════════════════════════════════
#  DATA
# ═══════════════════════════════════════════════════════════

def load_all(base):
    sd = os.path.dirname(os.path.abspath(__file__))
    data = {}
    for qn in Q_ORDER:
        qdir = os.path.join(base, qn)
        af = os.path.join(qdir, "analysis_results.json")
        sf = os.path.join(qdir, "sweep_results.json")
        swf = os.path.join(qdir, "switches.json")

        analysis = json.load(open(af)) if os.path.exists(af) else None
        sweep = json.load(open(sf)) if os.path.exists(sf) else None

        total_sw = 0
        if sweep: total_sw = sweep.get("num_switches", len(sweep.get("switches", [])))
        elif os.path.exists(swf): total_sw = len(json.load(open(swf)))

        if total_sw == 0 and analysis is None: continue

        sql_path = os.path.join(sd, "queries", f"{qn}.sql")
        sql = open(sql_path).read() if os.path.exists(sql_path) else ""

        sw_list = sweep.get("switches", []) if sweep else []

        data[qn] = {
            "analysis": analysis, "sweep": sweep,
            "results": analysis.get("results", []) if analysis else [],
            "total_switches": total_sw, "sql": sql, "sw_list": sw_list,
        }
    return data


# ═══════════════════════════════════════════════════════════
#  EXCEL — OVERVIEW
# ═══════════════════════════════════════════════════════════

def _build_overview(wb, data):
    ws = wb.active; ws.title = "Overview"; ws.sheet_properties.tabColor = "1E293B"
    ws.cell(row=1, column=1, value="TPC-H Query Optimizer — Switch Point Analysis Report").font = Font(bold=True, name="Arial", size=13, color="1E293B")
    ws.merge_cells("A1:K1")
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | TPC-H SF10 | PostgreSQL").font = Font(name="Arial", size=8, color="94A3B8")

    r = 4
    _hdr(ws, r, ["Query", "PSP Parameters", "Total\nSwitches", "Analyzed", "Skipped",
                  "Delayed", "Premature", "Correct", "Anomaly",
                  "Avg |Gap| (%)", "Max |Gap| (%)"])
    ws.row_dimensions[r].height = 32; r += 1

    for qn in Q_ORDER:
        if qn not in data: continue
        d = data[qn]; res = d["results"]
        nt = d["total_switches"]; na = len(res); ns = nt - na
        nd = sum(1 for x in res if x.get("direction") == "DELAYED")
        np_ = sum(1 for x in res if x.get("direction") == "PREMATURE")
        nc = sum(1 for x in res if x.get("direction") == "CORRECT")
        nA = sum(1 for x in res if x.get("direction") == "ANOMALY")
        abs_gaps = [abs(x.get("selectivity_gap_pct", 0)) for x in res if abs(x.get("selectivity_gap_pct", 0)) > 0.05]
        avg_g = sum(abs_gaps) / len(abs_gaps) if abs_gaps else 0
        max_g = max(abs_gaps) if abs_gaps else 0
        ps = ", ".join(PARAMS_MAP.get(qn, []))

        _rw(ws, r, [qn.upper(), ps, nt, na, ns, nd, np_, nc, nA, avg_g, max_g],
            [None, None, None, None, None, None, None, None, None, '0.00', '0.00'])
        ws.cell(row=r, column=1).font = BB
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="left")
        gc = ws.cell(row=r, column=11)
        if max_g > 10: gc.fill = RD
        elif max_g > 1: gc.fill = YL
        elif na > 0: gc.fill = GN
        r += 1

    _ww(ws, [10, 35, 10, 9, 9, 9, 10, 9, 9, 11, 11])


# ═══════════════════════════════════════════════════════════
#  EXCEL — PER-QUERY
# ═══════════════════════════════════════════════════════════

def _build_query_sheet(wb, qn, d):
    ws = wb.create_sheet(title=qn.upper())
    res = d["results"]; am = {x["switch_num"]: x for x in res}

    ws.cell(row=1, column=1, value=LABELS.get(qn, qn.upper())).font = Font(bold=True, name="Arial", size=12, color="1E293B")
    ws.merge_cells("A1:I1")
    ps = ", ".join(PARAMS_MAP.get(qn, []))
    ws.cell(row=2, column=1, value=f"Selectivity Predicates: {ps}").font = Font(name="Arial", size=9, color="64748B")

    r = 4; ws.cell(row=r, column=1, value="Query:").font = SF; r += 1
    for ln in d["sql"].strip().split("\n"):
        ws.cell(row=r, column=1, value=ln).font = SQ; r += 1

    # TABLE 1: All switches
    r += 2; ws.cell(row=r, column=1, value="All Planner Switch Points").font = TF; r += 1
    _hdr(ws, r, ["Switch #", "Sel (%)", "Param Values", "Plan A", "Plan B", "Cost A", "Cost B", "Direction", "Hints"])
    ws.row_dimensions[r].height = 28; r += 1
    for sw in d["sw_list"]:
        sn = sw.get("switch_num", 0); a = am.get(sn)
        sel = sw.get("selectivity_pct", sw.get("frac", 0) * 100)
        direction = a.get("direction", "?") if a else "SKIPPED"
        hints = ("✓/✓" if a and a.get("hint_a_verified") and a.get("hint_b_verified")
                 else "✓/✗" if a and a.get("hint_a_verified")
                 else "✗/✓" if a and a.get("hint_b_verified")
                 else "✗/✗" if a else "—")
        _rw(ws, r, [sn, sel, _pv(sw.get("param_values", [])),
                     sw.get("from_label", "")[:40], sw.get("to_label", "")[:40],
                     sw.get("from_cost", 0), sw.get("to_cost", 0), direction, hints],
            [None, '0.00', None, None, None, '#,##0', '#,##0', None, None])
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="left")
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="left")
        ws.cell(row=r, column=5).alignment = Alignment(horizontal="left")
        dc = ws.cell(row=r, column=8)
        fills = {"DELAYED": YL, "PREMATURE": BL, "CORRECT": GN, "ANOMALY": RD, "SKIPPED": GY}
        dc.fill = fills.get(direction, GY)
        r += 1

    # TABLE 2: Cross-force per analyzed switch
    for a in res:
        r += 2; sn = a["switch_num"]
        ws.cell(row=r, column=1, value=f"Switch #{sn} — Cross-Force ({a['direction']})").font = TF
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5); r += 1

        bs = a["before_switch"]; af = a["after_switch"]
        _hdr(ws, r, ["", f"Before Switch\nSel={bs['selectivity_pct']:.2f}%",
                      f"After Switch\nSel={af['selectivity_pct']:.2f}%", "Before Cost", "After Cost"])
        ws.row_dimensions[r].height = 30; r += 1

        _rw(ws, r, ["Normal (planner)", bs["normal_time"], af["normal_time"], bs["normal_cost"], af["normal_cost"]],
            [None, '#,##0.0', '#,##0.0', '#,##0', '#,##0'])
        ws.cell(row=r, column=1).font = BB; r += 1

        _rw(ws, r, ["Forced Plan A (P1)", bs["forced_p1_time"], af["forced_p1_time"], bs["forced_p1_cost"], af["forced_p1_cost"]],
            [None, '#,##0.0', '#,##0.0', '#,##0', '#,##0'])
        ws.cell(row=r, column=1).font = BB
        for c in range(1, 6): ws.cell(row=r, column=c).fill = P1F
        r += 1

        _rw(ws, r, ["Forced Plan B (P2)", bs["forced_p2_time"], af["forced_p2_time"], bs["forced_p2_cost"], af["forced_p2_cost"]],
            [None, '#,##0.0', '#,##0.0', '#,##0', '#,##0'])
        ws.cell(row=r, column=1).font = BB
        for c in range(1, 6): ws.cell(row=r, column=c).fill = P2F
        r += 1

        safe_div = lambda a, b: a / b if b > 0 else 0
        rb = safe_div(max(bs["forced_p1_time"], bs["forced_p2_time"]), min(bs["forced_p1_time"], bs["forced_p2_time"]))
        ra = safe_div(max(af["forced_p1_time"], af["forced_p2_time"]), min(af["forced_p1_time"], af["forced_p2_time"]))
        _rw(ws, r, ["Time Ratio", f"{rb:.2f}x", f"{ra:.2f}x", "", ""])
        ws.cell(row=r, column=1).font = SF
        for c in range(1, 4): ws.cell(row=r, column=c).fill = GY
        r += 1

        wb_ = "P2 ✓" if bs.get("p2_faster") else "P1 ✓"
        wa_ = "P1 ✓" if af.get("p1_faster") else "P2 ✓"
        _rw(ws, r, ["Faster Plan", wb_, wa_, "", ""])
        ws.cell(row=r, column=1).font = SF
        ws.cell(row=r, column=2).fill = YL if bs.get("p2_faster") else GN
        ws.cell(row=r, column=3).fill = YL if af.get("p1_faster") else GN
        r += 1

    # TABLE 3: True switch points
    if res:
        r += 2; ws.cell(row=r, column=1, value="True Switch Point Analysis").font = TF; r += 1
        _hdr(ws, r, ["Switch #", "Direction", "Planner\nSel (%)", "Planner Params",
                      "True\nSel (%)", "True Params", "Sel Diff (%)",
                      "P1 at True (ms)", "P2 at True (ms)"])
        ws.row_dimensions[r].height = 30; r += 1
        for a in res:
            gap = a.get("selectivity_gap_pct", 0)
            _rw(ws, r, [a["switch_num"], a["direction"],
                        a["planner_selectivity_pct"], _pv(a.get("planner_param_values", [])),
                        a["true_selectivity_pct"], _pv(a.get("true_param_values", [])),
                        gap, a.get("true_forced_p1_time", 0), a.get("true_forced_p2_time", 0)],
                [None, None, '0.00', None, '0.00', None, '+0.00', '#,##0.0', '#,##0.0'])
            ws.cell(row=r, column=4).alignment = Alignment(horizontal="left")
            ws.cell(row=r, column=6).alignment = Alignment(horizontal="left")
            fills_d = {"DELAYED": YL, "PREMATURE": BL, "CORRECT": GN, "ANOMALY": RD}
            ws.cell(row=r, column=2).fill = fills_d.get(a["direction"], GY)
            ag = abs(gap)
            gc = ws.cell(row=r, column=7)
            if ag > 10: gc.fill = RD
            elif ag > 1: gc.fill = YL
            else: gc.fill = GN
            r += 1

    _ww(ws, [12, 12, 14, 24, 14, 24, 12, 16, 16])


# ═══════════════════════════════════════════════════════════
#  PLOTS
# ═══════════════════════════════════════════════════════════

plt.rcParams.update({"font.family": "sans-serif", "font.size": 9})

def _ax(ax, title="", xl="", yl=""):
    ax.set_facecolor("#fafbfe")
    ax.set_title(title, fontweight="bold", color="#1e293b", pad=14, fontsize=12, loc="left")
    if xl: ax.set_xlabel(xl, color="#64748b", fontsize=10)
    if yl: ax.set_ylabel(yl, color="#64748b", fontsize=10)
    ax.tick_params(colors="#475569", labelsize=9)
    for s in ("top", "right"): ax.spines[s].set_visible(False)
    for s in ("bottom", "left"): ax.spines[s].set_color("#e2e8f0")
    ax.grid(axis="y", color="#f1f5f9", lw=0.8)
    ax.set_axisbelow(True)

def _sv(fig, path):
    fig.savefig(path, dpi=200, bbox_inches="tight", facecolor="white", pad_inches=0.2)
    plt.close(fig)


def build_plots(data, plotdir):
    os.makedirs(plotdir, exist_ok=True)
    qs = [q for q in Q_ORDER if q in data and data[q]["results"]]
    if not qs: return
    n = len(qs)

    # ── Collect per-query stats ──
    q_stats = {}
    for q in qs:
        res = data[q]["results"]
        abs_gaps = [abs(x.get("selectivity_gap_pct", 0)) for x in res]
        # Time ratios at switch: max(P1,P2)/min(P1,P2) at the after_switch point
        time_ratios = []
        for x in res:
            af = x.get("after_switch", {})
            p1, p2 = af.get("forced_p1_time", 0), af.get("forced_p2_time", 0)
            if p1 > 0 and p2 > 0:
                time_ratios.append(max(p1, p2) / min(p1, p2))
        q_stats[q] = {
            "max_gap": max(abs_gaps) if abs_gaps else 0,
            "avg_gap": sum(abs_gaps) / len(abs_gaps) if abs_gaps else 0,
            "min_tr": min(time_ratios) if time_ratios else 1,
            "avg_tr": sum(time_ratios) / len(time_ratios) if time_ratios else 1,
            "max_tr": max(time_ratios) if time_ratios else 1,
            "time_ratios": time_ratios,
            "abs_gaps": abs_gaps,
        }

    # ══════════════════════════════════════════════════
    # PLOT 1: Max Selectivity Gap per Query
    # ══════════════════════════════════════════════════
    fig, ax = plt.subplots(figsize=(10, 5.5))
    _ax(ax, "Maximum Selectivity Gap per Query", "Max |Selectivity Gap| (%)", "")

    mg = [q_stats[q]["max_gap"] for q in qs]
    y = range(n)
    bars = ax.barh(y, mg, color=[PAL[i % len(PAL)] for i in range(n)],
                   edgecolor="white", lw=0.5, height=0.6)
    ax.set_yticks(y); ax.set_yticklabels([q.upper() for q in qs])
    ax.invert_yaxis()
    for b, v in zip(bars, mg):
        if v > 0.05:
            ax.text(b.get_width() + 0.5, b.get_y() + b.get_height() / 2,
                    f"{v:.1f}%", va="center", fontsize=8, fontweight="bold", color="#1e293b")
    ax.axvline(5, color="#f59e0b", lw=1, ls="--", alpha=0.6)
    ax.text(5.3, -0.8, "5% threshold", fontsize=7, color="#f59e0b")

    # Formula annotation
    ax.text(0.98, 0.02, r"Gap = |sel$_{planner}$ − sel$_{true}$|",
            transform=ax.transAxes, fontsize=8, ha="right", va="bottom",
            color="#64748b", style="italic",
            bbox=dict(boxstyle="round,pad=0.3", facecolor="white", edgecolor="#e2e8f0"))
    _sv(fig, os.path.join(plotdir, "max_gap_per_query.png"))

    # ══════════════════════════════════════════════════
    # PLOT 2: Time Ratio at Switch Points (min/avg/max)
    # ══════════════════════════════════════════════════
    fig, ax = plt.subplots(figsize=(12, 5.5))
    _ax(ax, "Execution Time Ratio at Switch Points (Forced P1 vs P2)",
        "", "Time Ratio (max/min)")

    x = np.arange(n)
    mins = [q_stats[q]["min_tr"] for q in qs]
    avgs = [q_stats[q]["avg_tr"] for q in qs]
    maxs = [q_stats[q]["max_tr"] for q in qs]

    w = 0.25
    ax.bar(x - w, mins, w, label="Min", color="#93c5fd", edgecolor="white", lw=0.5)
    ax.bar(x, avgs, w, label="Avg", color="#3b82f6", edgecolor="white", lw=0.5)
    ax.bar(x + w, maxs, w, label="Max", color="#1e40af", edgecolor="white", lw=0.5)

    ax.set_xticks(x); ax.set_xticklabels([q.upper() for q in qs], rotation=45, ha="right")
    ax.axhline(1, color="#94a3b8", lw=0.8, ls="--")
    ax.legend(fontsize=8, framealpha=0.9)

    # Cap y-axis if outliers
    if max(maxs) > 20:
        ax.set_yscale("log")
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda y, _: f"{y:.0f}x" if y >= 1 else f"{y:.1f}x"))

    ax.text(0.98, 0.98, r"Ratio = $\frac{max(T_{P1}, T_{P2})}{min(T_{P1}, T_{P2})}$ at switch point",
            transform=ax.transAxes, fontsize=8, ha="right", va="top",
            color="#64748b", style="italic",
            bbox=dict(boxstyle="round,pad=0.3", facecolor="white", edgecolor="#e2e8f0"))
    _sv(fig, os.path.join(plotdir, "time_ratio_at_switch.png"))

    # ══════════════════════════════════════════════════
    # PLOT 3: Switch Point Accuracy Score per Query
    # ══════════════════════════════════════════════════
    fig, ax = plt.subplots(figsize=(10, 5.5))
    _ax(ax, "Switch Point Accuracy Score per Query", "", "Accuracy Score (0–1)")

    # Score = 1 - mean(|gap_i| / 100) for analyzed switches
    # 1.0 = perfect, 0.0 = worst
    scores = []
    for q in qs:
        gaps = q_stats[q]["abs_gaps"]
        if gaps:
            sc = 1.0 - sum(g / 100 for g in gaps) / len(gaps)
        else:
            sc = 1.0
        scores.append(max(0, sc))

    bars = ax.bar(range(n), scores, color=[PAL[i % len(PAL)] for i in range(n)],
                  edgecolor="white", lw=0.5, width=0.7)
    ax.set_xticks(range(n)); ax.set_xticklabels([q.upper() for q in qs], rotation=45, ha="right")
    ax.set_ylim(0, 1.05)
    ax.axhline(0.95, color="#22c55e", lw=1, ls="--", alpha=0.5)
    ax.text(n - 0.5, 0.955, "0.95 (good)", fontsize=7, color="#22c55e", ha="right")
    ax.axhline(0.90, color="#f59e0b", lw=1, ls="--", alpha=0.5)
    ax.text(n - 0.5, 0.905, "0.90 (fair)", fontsize=7, color="#f59e0b", ha="right")

    for b, s in zip(bars, scores):
        ax.text(b.get_x() + b.get_width() / 2, b.get_height() + 0.01,
                f"{s:.2f}", ha="center", fontsize=7, fontweight="bold", color="#1e293b")

    ax.text(0.98, 0.02,
            r"Score = 1 − $\frac{1}{N}\sum_{i=1}^{N} \frac{|gap_i|}{100}$"
            "\nN = analyzed switches per query",
            transform=ax.transAxes, fontsize=8, ha="right", va="bottom",
            color="#64748b", style="italic",
            bbox=dict(boxstyle="round,pad=0.4", facecolor="white", edgecolor="#e2e8f0"))
    _sv(fig, os.path.join(plotdir, "accuracy_score.png"))

    # ══════════════════════════════════════════════════
    # PLOT 4: Gap Distribution (box plot)
    # ══════════════════════════════════════════════════
    fig, ax = plt.subplots(figsize=(12, 5))
    _ax(ax, "Selectivity Gap Distribution per Query", "", "|Selectivity Gap| (%)")

    box_data = [q_stats[q]["abs_gaps"] for q in qs]
    bp = ax.boxplot(box_data, patch_artist=True, widths=0.5,
                    medianprops=dict(color="#1e293b", lw=1.5),
                    whiskerprops=dict(color="#94a3b8"),
                    capprops=dict(color="#94a3b8"),
                    flierprops=dict(marker="o", markersize=3, markerfacecolor="#ef4444", alpha=0.5))
    for patch, color in zip(bp["boxes"], [PAL[i % len(PAL)] for i in range(n)]):
        patch.set_facecolor(color); patch.set_alpha(0.6); patch.set_edgecolor("white")

    ax.set_xticks(range(1, n + 1)); ax.set_xticklabels([q.upper() for q in qs], rotation=45, ha="right")
    ax.axhline(5, color="#f59e0b", lw=1, ls="--", alpha=0.5)

    ax.text(0.98, 0.98, "Box: Q1–Q3, Whiskers: 1.5×IQR, Dots: outliers",
            transform=ax.transAxes, fontsize=7, ha="right", va="top",
            color="#94a3b8", style="italic")
    _sv(fig, os.path.join(plotdir, "gap_boxplot.png"))

    # ══════════════════════════════════════════════════
    # PLOT 5: Delayed vs Premature scatter (signed gap)
    # ══════════════════════════════════════════════════
    fig, ax = plt.subplots(figsize=(10, 6))
    _ax(ax, "Signed Selectivity Gap: Delayed (+) vs Premature (−)",
        "Planner Switch Selectivity (%)", "Selectivity Gap (%)")

    for qi, q in enumerate(qs):
        for x in data[q]["results"]:
            gap = x.get("selectivity_gap_pct", 0)
            sel = x.get("planner_selectivity_pct", 0)
            if abs(gap) < 0.05: continue
            color = "#f59e0b" if gap > 0 else "#3b82f6"
            ax.scatter(sel, gap, color=color, s=25, alpha=0.5, edgecolors="white", lw=0.3, zorder=3)

    ax.axhline(0, color="#1e293b", lw=0.8)
    ax.set_xlim(0, 100)

    # Add annotations
    ax.fill_between([0, 100], 0, 80, alpha=0.03, color="#f59e0b")
    ax.fill_between([0, 100], -80, 0, alpha=0.03, color="#3b82f6")
    ax.text(95, 5, "DELAYED\n(too late)", fontsize=9, ha="right", color="#f59e0b", fontweight="bold", alpha=0.6)
    ax.text(95, -8, "PREMATURE\n(too early)", fontsize=9, ha="right", color="#3b82f6", fontweight="bold", alpha=0.6)

    ax.text(0.02, 0.02, r"Gap = sel$_{planner}$ − sel$_{true}$"
            "\n+ = delayed, − = premature",
            transform=ax.transAxes, fontsize=8, ha="left", va="bottom",
            color="#64748b", style="italic",
            bbox=dict(boxstyle="round,pad=0.3", facecolor="white", edgecolor="#e2e8f0"))
    _sv(fig, os.path.join(plotdir, "delayed_vs_premature.png"))

    # ══════════════════════════════════════════════════
    # PLOT 6: Performance Penalty at Switch Points
    # ══════════════════════════════════════════════════
    fig, ax = plt.subplots(figsize=(10, 6))
    _ax(ax, "Performance Penalty vs Selectivity Gap",
        "|Selectivity Gap| (%)", "Time Ratio at Switch (worse/better plan)")

    for qi, q in enumerate(qs):
        for x in data[q]["results"]:
            gap = abs(x.get("selectivity_gap_pct", 0))
            af = x.get("after_switch", {})
            p1, p2 = af.get("forced_p1_time", 0), af.get("forced_p2_time", 0)
            if p1 > 0 and p2 > 0 and gap > 0.05:
                tr = max(p1, p2) / min(p1, p2)
                ax.scatter(gap, tr, color=PAL[qi % len(PAL)], s=30, alpha=0.6,
                           edgecolors="white", lw=0.3, zorder=3)

    ax.axhline(1, color="#94a3b8", lw=0.8, ls="--")
    ax.set_xlim(0, None)

    handles = [plt.Line2D([0], [0], marker="o", color="w", markerfacecolor=PAL[i % len(PAL)],
               markersize=5, label=q.upper()) for i, q in enumerate(qs)]
    ax.legend(handles=handles, loc="upper right", fontsize=6, ncol=3, framealpha=0.9)

    ax.text(0.02, 0.98, "Each dot = one switch point\n"
            r"Penalty = $\frac{max(T_{P1}, T_{P2})}{min(T_{P1}, T_{P2})}$",
            transform=ax.transAxes, fontsize=8, ha="left", va="top",
            color="#64748b", style="italic",
            bbox=dict(boxstyle="round,pad=0.3", facecolor="white", edgecolor="#e2e8f0"))
    _sv(fig, os.path.join(plotdir, "penalty_vs_gap.png"))

    print(f"  Plots → {plotdir}/")


# ═══════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════

def main():
    p = argparse.ArgumentParser()
    p.add_argument("--results", default="results")
    args = p.parse_args()
    base = os.path.join(os.path.dirname(os.path.abspath(__file__)), args.results)

    print(f"\n{'━'*50}\n  Report Builder\n{'━'*50}")
    data = load_all(base)
    print(f"  Loaded {len(data)} queries")

    wb = Workbook()
    _build_overview(wb, data)
    for qn in Q_ORDER:
        if qn not in data: continue
        _build_query_sheet(wb, qn, data[qn])
    outpath = os.path.join(base, "report.xlsx")
    wb.save(outpath)
    print(f"  Excel → {outpath}")

    build_plots(data, os.path.join(base, "plots"))
    print(f"{'━'*50}\n")


if __name__ == "__main__":
    main()