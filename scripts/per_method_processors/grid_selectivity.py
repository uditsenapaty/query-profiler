# ==========================================================
# scripts/per_method_processors/grid_selectivity.py
# ==========================================================
#
# Layout rules
# ------------
# 1D  : single column of cells (x1 values down rows)
# 2D  : plain grid  (x1 across cols, x2 down rows)
# 3D  : 2D blocks tiled HORIZONTALLY  for each x3 value
# 4D  : rows of 3D layouts stacked VERTICALLY  for each x4 value
# 5D  : columns of 4D layouts tiled HORIZONTALLY for each x5 value
# 6D  : rows of 5D layouts stacked VERTICALLY  for each x6 value
# ...  alternating H / V as depth increases
#
# GAP blank rows/cols separate every block group.
# Progressive spacing: x=1 gap after every 2-D block,
#                      x=2 gaps between x3 groups,
#                      x=3 gaps between x4 super-groups, etc.
# ==========================================================

from pathlib import Path
import pandas as pd
import numpy as np

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.formatting.rule import ColorScaleRule

# ----------------------------------------------------------
# colors / styles
# ----------------------------------------------------------

PINK       = PatternFill(fill_type="solid", fgColor="FFC0CB")
YELLOW     = PatternFill(fill_type="solid", fgColor="FFFF00")
BLUE       = PatternFill(fill_type="solid", fgColor="87CEFA")
RED_SIDE   = Side(style="thick", color="FF0000")
RED_BORDER = Border(
    left   = RED_SIDE,
    right  = RED_SIDE,
    top    = RED_SIDE,
    bottom = RED_SIDE,
)

BASE_GAP = 1  # blank rows/cols between adjacent 2D blocks


# ==========================================================
# helpers
# ==========================================================

def _discover_x_cols(df: pd.DataFrame) -> list:
    """Return sorted list of coordinate columns x1, x2, x3, ..."""
    cols = [c for c in df.columns if c.startswith("x") and c[1:].isdigit()]
    return sorted(cols, key=lambda c: int(c[1:]))


def _boundary_sets(df: pd.DataFrame, x_cols: list) -> dict:
    """
    For each coordinate column the boundary values are:
        min, second-smallest, max
    (matches the original 2-D logic, generalised to ND).
    """
    bsets = {}
    for c in x_cols:
        vals = np.sort(df[c].unique())
        s: set = set()
        if len(vals) >= 1:
            s.add(vals[0])
        if len(vals) >= 2:
            s.add(vals[1])
        if len(vals) >= 1:
            s.add(vals[-1])
        bsets[c] = s
    return bsets


def _is_boundary(row, x_cols: list, bsets: dict) -> bool:
    """Return True if ANY coordinate of this row is a boundary value."""
    for c in x_cols:
        if row[c] in bsets[c]:
            return True
    return False


def _all_ds_nonzero(row, ndim: int, axis: int) -> bool:
    """
    Red-border logic: True only when dS_x{k}_axis{axis} is
    non-zero and non-NaN for EVERY k in 1..ndim.
    """
    for k in range(1, ndim + 1):
        v = row.get(f"dS_x{k}_axis{axis}", np.nan)
        if pd.isna(v) or v == 0:
            return False
    return True


# ==========================================================
# ND layout engine
# ==========================================================

def _subtree_size(depth: int, extra_cols: list, sorted_vals: dict,
                  nx1: int, nx2: int) -> tuple:
    """
    Recursively compute (width_cols, height_rows) of the full
    sub-layout rooted at `depth` (0-based index into extra_cols).

    depth == len(extra_cols) means we are at a leaf 2-D block.

    Progressive gap: ((depth // 2) + 1) * BASE_GAP  between siblings at
    this level.
    """
    if depth == len(extra_cols):
        # leaf: one 2-D block
        return nx1+1, nx2+1

    col  = extra_cols[depth]
    n    = len(sorted_vals[col])
    gap = ((depth // 2) + 1) * BASE_GAP          # progressive spacing
    child_w, child_h = _subtree_size(
        depth + 1, extra_cols, sorted_vals, nx1, nx2
    )

    if depth % 2 == 0:
        # even depth (0, 2, 4, ...) → tile HORIZONTALLY (x3, x5, ...)
        total_w = n * child_w + (n - 1) * gap
        total_h = child_h
    else:
        # odd depth  (1, 3, 5, ...) → tile VERTICALLY  (x4, x6, ...)
        total_w = child_w
        total_h = n * child_h + (n - 1) * gap

    return total_w, total_h


def _iter_nd_blocks(sorted_vals: dict, x_cols: list):
    """
    Yield (combo_dict, row_offset, col_offset)  for every unique
    combination of x3, x4, x5, ... values.

    row_offset and col_offset are 0-based cell offsets from the
    top-left of the entire layout region (NOT including the axis
    headers; those are added by the caller).

    For 2D (no extra cols) yields one entry: ({}, 0, 0).
    """
    extra_cols = x_cols[2:]   # x3, x4, x5, ...

    nx1 = len(sorted_vals[x_cols[0]])
    nx2 = len(sorted_vals[x_cols[1]])

    def _recurse(depth: int, combo: dict, base_r: int, base_c: int):
        if depth == len(extra_cols):
            yield (combo, base_r, base_c)
            return

        col   = extra_cols[depth]
        vals  = sorted_vals[col]
        gap = ((depth // 2) + 1) * BASE_GAP
        child_w, child_h = _subtree_size(
            depth + 1, extra_cols, sorted_vals, nx1, nx2
        )

        for idx, val in enumerate(vals):
            new_combo = {**combo, col: val}
            if depth % 2 == 0:
                # tile horizontally
                c_off = base_c + idx * (child_w + gap)
                r_off = base_r
            else:
                # tile vertically
                r_off = base_r + idx * (child_h + gap)
                c_off = base_c

            yield from _recurse(depth + 1, new_combo, r_off, c_off)

    if not extra_cols:
        yield ({}, 0, 0)
    else:
        yield from _recurse(0, {}, 0, 0)


# ==========================================================
# 1-D sheet writer
# ==========================================================

def _create_sheet_1d(
    outfile,
    df: pd.DataFrame,
    x_cols: list,
    sorted_vals: dict,
    bsets: dict,
    value_col: str,
    axis: int,
    fill_color=None,
    use_qerr: bool = False,
    joint: bool = False,
):
    """Write a single-column grid for a 1-D dataset."""
    wb = Workbook()
    ws = wb.active

    x1   = x_cols[0]
    xvals = sorted_vals[x1]

    # header
    ws.cell(row=1, column=1, value=x1)
    ws.cell(row=1, column=2, value=value_col)

    sub_lookup = {
        r[x1]: r
        for _, r in df.iterrows()
    }

    qerr_cells = []

    for i, x in enumerate(xvals):
        excel_row = i + 2   # row 1 = header

        ws.cell(row=excel_row, column=1, value=round(float(x), 6))

        if x not in sub_lookup:
            continue

        r   = sub_lookup[x]
        val = r.get(value_col, np.nan)
        if isinstance(val, float) and np.isnan(val):
            val = None

        cell = ws.cell(row=excel_row, column=2, value=val)

        boundary = _is_boundary(r, x_cols, bsets)

        if boundary:
            cell.fill = PINK
        elif fill_color is not None and val is not None and val != 0:
            cell.fill = fill_color

        if joint and not boundary:
            if _all_ds_nonzero(r, len(x_cols), axis):
                cell.border = RED_BORDER

        if use_qerr and not boundary:
            q = r.get(f"adjacent_qerr_x{axis}", np.nan)
            if pd.notna(q):
                cell.value = q
                qerr_cells.append(cell.coordinate)

    if qerr_cells:
        for coord in qerr_cells:
            ws.conditional_formatting.add(
                coord,
                ColorScaleRule(
                    start_type="min", start_color="FFFFFF",
                    end_type="max",   end_color="006400",
                ),
            )

    wb.save(outfile)


# ==========================================================
# ND sheet writer  (2D and above)
# ==========================================================

def _create_sheet(
    outfile,
    df: pd.DataFrame,
    x_cols: list,
    sorted_vals: dict,
    bsets: dict,
    value_col: str,
    axis: int,
    fill_color=None,
    use_qerr: bool = False,
    joint: bool = False,
):
    wb = Workbook()
    ws = wb.active

    ndim  = len(x_cols)
    x1    = x_cols[0]
    x2    = x_cols[1]
    xvals = sorted_vals[x1]
    yvals = sorted_vals[x2]
    nx1   = len(xvals)
    nx2   = len(yvals)

    # Each block occupies:
    #   1 header row  (x1 values)
    #   nx2 data rows
    #   1 header col  (x2 values)
    #   nx1 data cols
    # The layout engine returns 0-based DATA offsets.
    # We add +1 for the header row, +1 for the header col,
    # then +1 for Excel 1-based indexing = +2 total shift.

    HDR_ROW_OFF = 1   # rows consumed by x1 axis header above each block
    HDR_COL_OFF = 1   # cols consumed by x2 axis header left of each block

    qerr_cells = []

    for combo, row_off, col_off in _iter_nd_blocks(sorted_vals, x_cols):

        # Filter df to this combo slice
        mask = pd.Series(True, index=df.index)
        for c, v in combo.items():
            mask &= (df[c] == v)
        sub_df = df[mask]

        sub_lookup = {
            (r[x1], r[x2]): r
            for _, r in sub_df.iterrows()
        }

        # Excel 1-based origin of this block's top-left corner
        # (the corner that holds the label / x2-header column / x1-header row)
        origin_r = row_off + 1   # 1-based
        origin_c = col_off + 1

        # ---- x1 axis header (top row of this block) ----
        for j, x in enumerate(xvals):
            ws.cell(
                row    = origin_r,
                column = origin_c + HDR_COL_OFF + j,
                value  = round(float(x), 6),
            )

        # ---- x2 axis header (left col of this block) ----
        for i, y in enumerate(yvals):
            ws.cell(
                row    = origin_r + HDR_ROW_OFF + i,
                column = origin_c,
                value  = round(float(y), 6),
            )

        # ---- group label in top-left corner cell ----
        if combo:
            # Only show labels at the boundary of each dimension group
            # (the top-left corner of every block always shows the combo)
            label = "  ".join(f"{c}={v}" for c, v in combo.items())
            ws.cell(row=origin_r, column=origin_c, value=label)

        # ---- data cells ----
        for i, y in enumerate(yvals):
            for j, x in enumerate(xvals):

                key = (x, y)
                if key not in sub_lookup:
                    continue

                r   = sub_lookup[key]
                val = r.get(value_col, np.nan)
                if isinstance(val, float) and np.isnan(val):
                    val = None

                cell = ws.cell(
                    row    = origin_r + HDR_ROW_OFF + i,
                    column = origin_c + HDR_COL_OFF + j,
                    value  = val,
                )

                boundary = _is_boundary(r, x_cols, bsets)

                if boundary:
                    cell.fill = PINK
                elif fill_color is not None and val is not None and val != 0:
                    cell.fill = fill_color

                if joint and not boundary:
                    if _all_ds_nonzero(r, ndim, axis):
                        cell.border = RED_BORDER

                if use_qerr and not boundary:
                    q = r.get(f"adjacent_qerr_x{axis}", np.nan)
                    if pd.notna(q):
                        cell.value = q
                        qerr_cells.append(cell.coordinate)

    # ---- qerr conditional formatting (one rule per cell) ----
    if qerr_cells:
        for coord in qerr_cells:
            ws.conditional_formatting.add(
                coord,
                ColorScaleRule(
                    start_type="min", start_color="FFFFFF",
                    end_type="max",   end_color="006400",
                ),
            )

    wb.save(outfile)


# ==========================================================
# public entry point
# ==========================================================

def run(results_dir):

    results_dir = Path(results_dir)
    csv_path    = results_dir / "ground_truth.csv"

    if not csv_path.exists():
        return

    df = pd.read_csv(csv_path)

    outdir = results_dir / "grid_selectivity"
    outdir.mkdir(parents=True, exist_ok=True)

    # ----------------------------------------------------------
    # discover coordinate columns dynamically
    # ----------------------------------------------------------
    x_cols = _discover_x_cols(df)
    ndim   = len(x_cols)

    if ndim < 1:
        print("No coordinate columns (x1, x2, ...) found. Skipping.")
        return

    # ----------------------------------------------------------
    # sorted unique values per coordinate
    # ----------------------------------------------------------
    sorted_vals = {c: np.sort(df[c].unique()) for c in x_cols}

    # ----------------------------------------------------------
    # boundary sets
    # ----------------------------------------------------------
    bsets = _boundary_sets(df, x_cols)

    # ----------------------------------------------------------
    # choose writer based on dimensionality
    # ----------------------------------------------------------
    writer_fn = _create_sheet_1d if ndim == 1 else _create_sheet

    # ----------------------------------------------------------
    # generate one workbook per (axis, param) pair
    # plus one joint workbook per axis
    # ----------------------------------------------------------
    for axis in range(1, ndim + 1):

        for param in range(1, ndim + 1):

            col = f"dS_x{param}_axis{axis}"     # FIX: don't use x_cols[param-1]

            if col not in df.columns:
                continue

            # x1-param -> YELLOW, all others -> BLUE
            fill = YELLOW if param == 1 else BLUE

            writer_fn(
                outfile     = outdir / f"axis{axis}_dS_x{param}.xlsx",
                df          = df,
                x_cols      = x_cols,
                sorted_vals = sorted_vals,
                bsets       = bsets,
                value_col   = col,
                axis        = axis,
                fill_color  = fill,
            )

        joint_col = f"joint_dS_axis{axis}"

        if joint_col in df.columns:
            writer_fn(
                outfile     = outdir / f"axis{axis}_joint_dS.xlsx",
                df          = df,
                x_cols      = x_cols,
                sorted_vals = sorted_vals,
                bsets       = bsets,
                value_col   = joint_col,
                axis        = axis,
                use_qerr    = True,
                joint       = True,
            )

    print(f"saved: {outdir}")