# ==========================================================
# scripts/per_method_processors/grid_qerr.py
# ==========================================================

from pathlib import Path

import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side

from openpyxl.formatting.rule import (
    ColorScaleRule
)

# ==========================================================
# colors
# ==========================================================

PINK = PatternFill(
    fill_type="solid",
    fgColor="FFC0CB"
)

CYAN = PatternFill(
    fill_type="solid",
    fgColor="00FFFF"
)

RED_BORDER = Border(
    left=Side(style="thick", color="FF0000"),
    right=Side(style="thick", color="FF0000"),
    top=Side(style="thick", color="FF0000"),
    bottom=Side(style="thick", color="FF0000"),
)

GAP = 1


# ==========================================================
# helpers
# ==========================================================

def discover_x_cols(df):

    cols = [

        c

        for c in df.columns

        if (
            c.startswith("x")
            and
            c[1:].isdigit()
        )
    ]

    return sorted(
        cols,
        key=lambda c: int(c[1:])
    )


def boundary_sets(df, x_cols):

    out = {}

    for c in x_cols:

        vals = np.sort(
            df[c].unique()
        )

        s = set()

        if len(vals) >= 1:
            s.add(vals[0])

        if len(vals) >= 2:
            s.add(vals[1])

        if len(vals) >= 1:
            s.add(vals[-1])

        out[c] = s

    return out


def is_boundary(row, x_cols, bsets):

    for c in x_cols:

        if row[c] in bsets[c]:
            return True

    return False


# ==========================================================
# zero-row-count detection
# ==========================================================

def _is_zero(v):

    try:
        return float(v) == 0.0

    except (TypeError, ValueError):
        return False


def build_count_lookup(df, x_cols):

    if "count_rows" not in df.columns:
        return {}

    lookup = {}

    for _, r in df.iterrows():

        key = tuple(
            float(r[c])
            for c in x_cols
        )

        lookup[key] = r["count_rows"]

    return lookup


def has_zero_count(
    row,
    axis,
    x_cols,
    count_lookup
):

    # this cell's own row count
    if _is_zero(
        row.get("count_rows")
    ):
        return True

    # the neighbour used for this axis' adjacent qerr
    ncols = [
        f"x{d}_neighbor_axis{axis}"
        for d in range(1, len(x_cols) + 1)
    ]

    if not all(
        c in row.index
        for c in ncols
    ):
        return False

    raw = [row[c] for c in ncols]

    if any(pd.isna(v) for v in raw):
        return False

    ncoords = tuple(
        float(v) for v in raw
    )

    neigh = count_lookup.get(
        ncoords
    )

    return _is_zero(neigh)


def pick_fill(
    row,
    axis,
    x_cols,
    bsets,
    count_lookup
):

    # cyan (a zero row count on this cell or its neighbour)
    # overlays the pink boundary fill when both apply
    if has_zero_count(
        row,
        axis,
        x_cols,
        count_lookup
    ):
        return CYAN

    if is_boundary(
        row,
        x_cols,
        bsets
    ):
        return PINK

    return None


# ==========================================================
# recursive ND placement
# ==========================================================

def build_layout(
    sorted_vals,
    extra_cols,
    depth=0
):

    if len(extra_cols) == 0:

        return [
            ({}, 0, 0)
        ]

    first = extra_cols[0]

    rest = extra_cols[1:]

    child = build_layout(
        sorted_vals,
        rest,
        depth + 1
    )

    placements = []

    horizontal = (
        depth % 2 == 0
    )

    max_row = max(
        r for _, r, _ in child
    )

    max_col = max(
        c for _, _, c in child
    )

    child_h = max_row + 1
    child_w = max_col + 1

    for idx, val in enumerate(
        sorted_vals[first]
    ):

        for combo, r, c in child:

            combo2 = dict(combo)

            combo2[first] = val

            if horizontal:

                placements.append(
                    (
                        combo2,
                        r,
                        c + idx * child_w
                    )
                )

            else:

                placements.append(
                    (
                        combo2,
                        r + idx * child_h,
                        c
                    )
                )

    return placements


# ==========================================================
# per-axis sheet writer
# ==========================================================

def write_axis_sheet(

    ws,
    df,
    axis,
    x_cols,
    bsets,
    count_lookup

):

    ndim = len(x_cols)

    qcol = (
        f"adjacent_qerr_x{axis}"
    )

    plan_col = (
        f"plan_change_x{axis}"
    )

    if qcol not in df.columns:
        return

    # ======================================================
    # 1D
    # ======================================================

    if ndim == 1:

        x = x_cols[0]

        vals = np.sort(
            df[x].unique()
        )

        qcells = []

        ws.cell(
            row=1,
            column=1,
            value=x
        )

        ws.cell(
            row=1,
            column=2,
            value="qerr"
        )

        for i, xv in enumerate(
            vals,
            start=2
        ):

            row = df.loc[
                df[x] == xv
            ].iloc[0]

            ws.cell(
                row=i,
                column=1,
                value=float(xv)
            )

            cell = ws.cell(
                row=i,
                column=2,
                value=row[qcol]
            )

            if (
                plan_col in df.columns
                and bool(row[plan_col])
            ):
                cell.border = RED_BORDER

            fill = pick_fill(
                row,
                axis,
                x_cols,
                bsets,
                count_lookup
            )

            if fill is not None:

                cell.fill = fill

            else:

                qcells.append(
                    cell.coordinate
                )

        if qcells:

            ws.conditional_formatting.add(

                " ".join(qcells),

                ColorScaleRule(

                    start_type="min",
                    start_color="FFFFFF",

                    end_type="max",
                    end_color="006400"

                )
            )

        return

    # ======================================================
    # ND
    # ======================================================

    sorted_vals = {

        c: np.sort(
            df[c].unique()
        )

        for c in x_cols
    }

    x1 = x_cols[0]
    x2 = x_cols[1]

    xvals = sorted_vals[x1]
    yvals = sorted_vals[x2]

    extra_cols = x_cols[2:]

    placements = build_layout(
        sorted_vals,
        extra_cols
    )

    block_w = len(xvals) + 2 + GAP
    block_h = len(yvals) + 2 + GAP

    qcells = []

    for combo, br, bc in placements:

        mask = np.ones(
            len(df),
            dtype=bool
        )

        for c, v in combo.items():

            mask &= (
                df[c] == v
            )

        sub = df.loc[mask]

        lookup = {

            (r[x1], r[x2]): r

            for _, r in sub.iterrows()
        }

        start_row = (
            br * block_h
        ) + 1

        start_col = (
            bc * block_w
        ) + 1

        label = ", ".join(

            f"{k}={v}"

            for k, v in combo.items()
        )

        ws.cell(
            row=start_row,
            column=start_col,
            value=label
        )

        for j, xv in enumerate(
            xvals,
            start=2
        ):

            ws.cell(
                row=start_row,
                column=start_col + j - 1,
                value=float(xv)
            )

        for i, yv in enumerate(
            yvals,
            start=2
        ):

            ws.cell(
                row=start_row + i - 1,
                column=start_col,
                value=float(yv)
            )

        for i, yv in enumerate(yvals):

            for j, xv in enumerate(xvals):

                row = lookup.get(
                    (xv, yv)
                )

                if row is None:
                    continue

                cell = ws.cell(

                    row=start_row + i + 1,

                    column=start_col + j + 1,

                    value=row[qcol]

                )

                if (
                    plan_col in df.columns
                    and bool(row[plan_col])
                ):
                    cell.border = RED_BORDER

                fill = pick_fill(
                    row,
                    axis,
                    x_cols,
                    bsets,
                    count_lookup
                )

                if fill is not None:

                    cell.fill = fill

                else:

                    qcells.append(
                        cell.coordinate
                    )

    if qcells:

        ws.conditional_formatting.add(

            " ".join(qcells),

            ColorScaleRule(

                start_type="min",
                start_color="FFFFFF",

                end_type="max",
                end_color="006400"

            )
        )


# ==========================================================
# entry
# ==========================================================

def run(results_dir):

    results_dir = Path(
        results_dir
    )

    csv = (
        results_dir
        / "ground_truth.csv"
    )

    if not csv.exists():
        return

    df = pd.read_csv(csv)

    x_cols = discover_x_cols(
        df
    )

    ndim = len(x_cols)

    if ndim == 0:
        return

    bsets = boundary_sets(
        df,
        x_cols
    )

    count_lookup = build_count_lookup(
        df,
        x_cols
    )

    outfile = (
        results_dir
        / "grid_qerr.xlsx"
    )

    wb = Workbook()

    default_ws = wb.active

    written = 0

    for axis in range(
        1,
        ndim + 1
    ):

        qcol = (
            f"adjacent_qerr_x{axis}"
        )

        if qcol not in df.columns:
            continue

        ws = wb.create_sheet(
            title=f"axis{axis}"
        )

        write_axis_sheet(
            ws,
            df,
            axis,
            x_cols,
            bsets,
            count_lookup
        )

        written += 1

    if written == 0:
        return

    wb.remove(
        default_ws
    )

    wb.save(outfile)

    print(
        f"saved: {outfile}"
    )

if __name__=="__main__":

    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "gt_results_sf1_10x10_s1q0/qt8/m0"
    run(path)
