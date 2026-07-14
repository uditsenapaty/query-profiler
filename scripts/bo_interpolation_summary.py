#!/usr/bin/env python3
# =========================================================
# scripts/bo_interpolation_summary.py
# =========================================================
# A nice tabbed .xlsx summary + conclusion of the BO-interpolation
# experiment over a whole gt_results_* root.
#
# Reads   {gt_root}/bo_interpolation_eval/bo_eval_all.csv
#          (produced by evaluate_bo_interpolation.py)
# Writes  {gt_root}/bo_interpolation_summary.xlsx  with tabs:
#          Conclusion | Ranking | By Representation | By Kernel | Detail
#
# Resolution / default root are taken from config_gt, so this can be
# run at the tail of a run with no arguments.
#
# Usage:
#   python scripts/bo_interpolation_summary.py                       # root from config
#   python scripts/bo_interpolation_summary.py gt_results_sf1_10x10_s1q0
# =========================================================

import sys
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config_gt


BASELINES = {"random", "uniform_stride"}
PRIMARY   = "QERR_MEDIAN_UNSEEN"          # lower = better interpolation accuracy
DISCOVERY = "max_found_pct"               # higher = better boundary discovery

AGG_COLS = [
    "QERR_MEDIAN_UNSEEN", "QERR_P90_UNSEEN", "QERR_MAX_UNSEEN",
    "WITHIN_2X_UNSEEN", "WITHIN_5X_UNSEEN", "RMSE_UNSEEN", "R2_UNSEEN",
    "max_found_pct", "coverage_2x_pct", "coverage_5x_pct",
    "S_max_relerr", "S_avg_relerr", "S_topk_relerr",
]

# --------------------------------------------------------- styling
C_TITLE = "1F3864"; C_HDR = "4472C4"; C_SECTION = "8EAADB"
C_BAND  = "EDF2FB"; C_GOOD = "C6EFCE"; WHITE = "FFFFFF"
_THIN   = Side(style="thin", color="BFBFBF")
BORDER  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _fill(c):
    return PatternFill("solid", fgColor=c)


def _title(ws, row, ncols, text, color=C_TITLE, size=14, h=26):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color=WHITE, size=size)
    c.fill = _fill(color)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = h
    return row + 1


def _num(v):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return None, None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return str(v), None
    if float(f).is_integer() and abs(f) < 1e15:
        return int(f), "#,##0"
    return f, "0.000"


def _df_sheet(ws, df, title, highlight_first=False):
    headers = list(df.columns)
    ncols = max(len(headers), 1)

    r = _title(ws, 1, ncols, title)
    r += 1

    for j, name in enumerate(headers, start=1):
        c = ws.cell(row=r, column=j, value=str(name))
        c.font = Font(bold=True, color=WHITE, size=9)
        c.fill = _fill(C_HDR)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[r].height = 26
    r += 1

    band = _fill(C_BAND); good = _fill(C_GOOD)
    for i, (_, row) in enumerate(df.iterrows()):
        for j, name in enumerate(headers, start=1):
            val, nf = _num(row[name])
            c = ws.cell(row=r, column=j, value=val)
            c.border = BORDER
            c.font = Font(size=9)
            if nf:
                c.number_format = nf
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
            if highlight_first and i == 0:
                c.fill = good
            elif i % 2 == 1:
                c.fill = band
        r += 1

    ws.column_dimensions["A"].width = 22
    for j in range(2, ncols + 1):
        w = max(9, min(20, max(len(str(headers[j - 1])) + 2,
                               int(df.iloc[:, j - 1].astype(str).map(len).max()) + 1)))
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(row=4, column=1)


def _conclusion_sheet(ws, info, kv, bullets):
    ncols = 6
    r = _title(ws, 1, ncols, "BO INTERPOLATION — CONCLUSION", size=15, h=28)
    r += 1

    for label, value in info:
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(bold=True, size=10, color="404040")
        lc.fill = _fill("D9E1F2")
        lc.border = BORDER
        lc.alignment = Alignment(horizontal="left", indent=1, vertical="center")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=ncols)
        vc = ws.cell(row=r, column=2, value=str(value))
        vc.font = Font(size=10); vc.border = BORDER
        vc.alignment = Alignment(horizontal="left", indent=1, vertical="center")
        r += 1

    r += 1
    r = _title(ws, r, ncols, "Key results", color=C_SECTION, size=12, h=20)
    for label, value in kv:
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(bold=True, size=10)
        lc.border = BORDER
        lc.alignment = Alignment(horizontal="left", indent=1, vertical="center")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=ncols)
        vc = ws.cell(row=r, column=2, value=str(value))
        vc.font = Font(size=10); vc.border = BORDER
        vc.alignment = Alignment(horizontal="left", indent=1, vertical="center")
        r += 1

    r += 1
    r = _title(ws, r, ncols, "Conclusion", color=C_SECTION, size=12, h=20)
    for line in bullets:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        c = ws.cell(row=r, column=1, value="•  " + line)
        c.font = Font(size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        ws.row_dimensions[r].height = 30
        r += 1

    ws.column_dimensions["A"].width = 30
    for j in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(j)].width = 22
    ws.sheet_view.showGridLines = False


# --------------------------------------------------------- aggregation
def _agg_by(df, by):
    cols = [c for c in AGG_COLS if c in df.columns]
    g = df.groupby(by, as_index=False)[cols].mean(numeric_only=True)
    return g


def _pct_better(base_v, new_v):
    if base_v in (0, None) or np.isnan(base_v) or np.isnan(new_v):
        return float("nan")
    return (base_v - new_v) / base_v * 100.0


# --------------------------------------------------------- main
def run(gt_root):
    gt_root = Path(gt_root)
    csv = gt_root / "bo_interpolation_eval" / "bo_eval_all.csv"
    if not csv.exists():
        print(f"Missing {csv} — run evaluate_bo_interpolation.py first.")
        return

    df = pd.read_csv(csv)
    df_noref = df[df["method"] != "ground_truth"].copy()

    # per-method aggregate across (query × gt_method)
    keep = [c for c in AGG_COLS if c in df_noref.columns]
    agg = df_noref.groupby("method").agg(
        representation=("representation", "first"),
        kernel=("kernel", "first"),
        acquisition=("acquisition", "first"),
        **{c: (c, "mean") for c in keep},
        n_evals=("method", "size"),
    ).reset_index()

    agg = agg.sort_values(PRIMARY).reset_index(drop=True)
    agg.insert(0, "rank", agg.index + 1)

    bo_agg   = agg[agg["method"].str.startswith("bo_")].reset_index(drop=True)
    base_agg = agg[agg["method"].isin(BASELINES)].reset_index(drop=True)

    best_overall = agg.iloc[0]
    best_bo      = bo_agg.iloc[0] if len(bo_agg) else None
    best_base    = base_agg.sort_values(PRIMARY).iloc[0] if len(base_agg) else None

    bo_rows = df_noref[df_noref["method"].str.startswith("bo_")]
    rep_agg  = _agg_by(bo_rows, "representation").sort_values(PRIMARY).reset_index(drop=True)
    kern_agg = _agg_by(bo_rows, "kernel").sort_values(PRIMARY).reset_index(drop=True)
    acq_agg  = _agg_by(bo_rows, "acquisition").sort_values(PRIMARY).reset_index(drop=True)

    disc = agg.sort_values(DISCOVERY, ascending=False).reset_index(drop=True)
    best_disc = disc.iloc[0]

    # ---- conclusion text ----
    n_q = df["query"].nunique() if "query" in df.columns else 0
    n_m = df.groupby(["query", "gt_method"]).ngroups if "gt_method" in df.columns else 0
    budget_pct = 100.0 * float(df_noref["sample_fraction"].mean()) if "sample_fraction" in df else float("nan")

    info = [
        ("Folder", gt_root.name),
        ("Queries", n_q),
        ("Method dirs evaluated", n_m),
        ("Budget (mean sampled)", f"{budget_pct:.1f}% of pairs"),
        ("Representations", "A = pair (P1,P2) [default]   |   B = point P (deterministic neighbour)"),
        ("Kernels", ", ".join(sorted(bo_rows['kernel'].unique())) if len(bo_rows) else "-"),
        ("Acquisitions", ", ".join(sorted(bo_rows['acquisition'].unique())) if len(bo_rows) else "-"),
        ("Primary metric", "QERR_MEDIAN_UNSEEN  (pred q-error on UNSEEN pairs; lower is better)"),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]

    def _lbl(m):
        rep = agg.loc[agg.method == m, "representation"].iloc[0]
        ker = agg.loc[agg.method == m, "kernel"].iloc[0]
        acq = agg.loc[agg.method == m, "acquisition"].iloc[0]
        return f"{m}" + (f"  (rep={rep}, kernel={ker}, acq={acq})" if rep != "-" else "")

    kv = [
        ("Best overall", f"{_lbl(best_overall['method'])}  →  median unseen q-error "
                         f"{best_overall[PRIMARY]:.3f}"),
    ]
    if best_bo is not None:
        kv.append(("Best BO", f"{_lbl(best_bo['method'])}  →  {best_bo[PRIMARY]:.3f}"))
    if best_base is not None:
        kv.append(("Best baseline", f"{best_base['method']}  →  {best_base[PRIMARY]:.3f}"))
    if len(rep_agg):
        kv.append(("Representation means",
                   "   ".join(f"{r['representation']}={r[PRIMARY]:.3f}"
                             for _, r in rep_agg.iterrows())))
    if len(kern_agg):
        kv.append(("Best kernel",
                   f"{kern_agg.iloc[0]['kernel']}  →  {kern_agg.iloc[0][PRIMARY]:.3f}"))
    if len(acq_agg):
        kv.append(("Best acquisition",
                   f"{acq_agg.iloc[0]['acquisition']}  →  {acq_agg.iloc[0][PRIMARY]:.3f}"))
    kv.append(("Best boundary discovery",
               f"{_lbl(best_disc['method'])}  →  finds {best_disc[DISCOVERY]:.0f}% of true max qerr, "
               f"coverage>5x {best_disc.get('coverage_5x_pct', float('nan')):.0f}%"))

    bullets = []
    if best_bo is not None and best_base is not None:
        d = _pct_better(best_base[PRIMARY], best_bo[PRIMARY])
        if not np.isnan(d):
            if d > 1:
                bullets.append(f"BO improves median unseen q-error by {d:.1f}% vs the best "
                               f"baseline ({best_base['method']}): {best_bo[PRIMARY]:.3f} vs "
                               f"{best_base[PRIMARY]:.3f}.")
            elif d < -1:
                bullets.append(f"The baseline ({best_base['method']}) is competitive: BO is "
                               f"{-d:.1f}% worse on median unseen q-error here.")
            else:
                bullets.append(f"BO and the best baseline are within 1% on median unseen q-error "
                               f"({best_bo[PRIMARY]:.3f} vs {best_base[PRIMARY]:.3f}).")

    if len(rep_agg) == 2:
        a = rep_agg[rep_agg.representation == "pair"][PRIMARY]
        b = rep_agg[rep_agg.representation == "point"][PRIMARY]
        if len(a) and len(b):
            av, bv = float(a.iloc[0]), float(b.iloc[0])
            diff = _pct_better(max(av, bv), min(av, bv))
            better = "pair (A)" if av < bv else "point (B)"
            if abs(av - bv) / max(av, bv) < 0.05:
                bullets.append(f"Representations are comparable (pair={av:.3f}, point={bv:.3f}, "
                               f"<5% apart) → prefer the lower-dimensional point (B) representation.")
            else:
                bullets.append(f"{better} representation is better ({av:.3f} vs {bv:.3f}, "
                               f"{diff:.1f}% gap).")

    if len(kern_agg):
        bullets.append(f"Best kernel: {kern_agg.iloc[0]['kernel']} "
                       f"(median unseen q-error {kern_agg.iloc[0][PRIMARY]:.3f}); "
                       f"worst: {kern_agg.iloc[-1]['kernel']} ({kern_agg.iloc[-1][PRIMARY]:.3f}).")
    if len(acq_agg):
        bullets.append(f"Best acquisition: {acq_agg.iloc[0]['acquisition']} "
                       f"({acq_agg.iloc[0][PRIMARY]:.3f}); worst: "
                       f"{acq_agg.iloc[-1]['acquisition']} ({acq_agg.iloc[-1][PRIMARY]:.3f}).")

    bullets.append(f"For finding boundary (high-qerr) pairs, {best_disc['method']} recovered "
                   f"{best_disc[DISCOVERY]:.0f}% of the true max q-error within the "
                   f"{budget_pct:.0f}% budget.")

    # ---- workbook ----
    wb = Workbook(); wb.remove(wb.active)

    _conclusion_sheet(wb.create_sheet("Conclusion"), info, kv, bullets)

    rank_cols = ["rank", "method", "representation", "kernel", "acquisition", PRIMARY,
                 "QERR_P90_UNSEEN", "WITHIN_2X_UNSEEN", "WITHIN_5X_UNSEEN",
                 "RMSE_UNSEEN", "R2_UNSEEN", "max_found_pct",
                 "coverage_5x_pct", "S_max_relerr", "n_evals"]
    rank_cols = [c for c in rank_cols if c in agg.columns]
    _df_sheet(wb.create_sheet("Ranking"), agg[rank_cols],
              "Ranked methods (mean over queries × method dirs) — sorted by median unseen q-error",
              highlight_first=True)

    if len(rep_agg):
        _df_sheet(wb.create_sheet("By Representation"), rep_agg,
                  "Representation comparison (BO only, mean over kernels/queries)")
    if len(kern_agg):
        _df_sheet(wb.create_sheet("By Kernel"), kern_agg,
                  "Kernel comparison (BO only, mean over representations/queries)")
    if len(acq_agg):
        _df_sheet(wb.create_sheet("By Acquisition"), acq_agg,
                  "Acquisition comparison (BO only, mean over kernels/representations/queries)")

    detail_cols = ["query", "gt_method", "method", "representation", "kernel", "acquisition",
                   PRIMARY, "QERR_P90_UNSEEN", "WITHIN_2X_UNSEEN",
                   "max_qerr_found", "max_found_pct", "coverage_5x_pct", "budget"]
    detail_cols = [c for c in detail_cols if c in df.columns]
    detail = df[detail_cols].sort_values(
        [c for c in ["query", "gt_method", PRIMARY] if c in detail_cols]
    ).reset_index(drop=True)
    _df_sheet(wb.create_sheet("Detail"), detail,
              "Per query × method-dir × BO-method detail")

    out = gt_root / "bo_interpolation_summary.xlsx"
    wb.save(out)

    print(f"\nSaved: {out}")
    print(f"  Best overall : {best_overall['method']}  ({PRIMARY}={best_overall[PRIMARY]:.3f})")
    if best_bo is not None:
        print(f"  Best BO      : {best_bo['method']}  ({best_bo[PRIMARY]:.3f})")
    if len(kern_agg):
        print(f"  Best kernel  : {kern_agg.iloc[0]['kernel']}")
    return out


def _default_root():
    res = config_gt.get_query_resolution(config_gt.QUERIES[0], config_gt.RUN_METHODS[0])
    return config_gt.get_main_dir(res)


if __name__ == "__main__":
    arg = sys.argv[1] if len(sys.argv) > 1 else None
    run(Path(arg) if arg else _default_root())
