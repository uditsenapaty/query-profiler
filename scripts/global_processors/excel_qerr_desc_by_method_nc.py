# =========================================================
# scripts/excel_qerr_desc_by_method_nc.py
#
# Like excel_qerr_desc_nc.py, but keeps each method in its OWN
# workbook instead of stacking all methods inside one tab:
#
#   - one .xlsx per method (m0, m1, m2)
#   - each workbook has 5 tabs, one per query (qt5, qt7, qt8,
#     qt10, qt16), showing ONLY that method's qerr_desc_nc detail
#
# Only _nc (non-zero cardinality) data is used. Rows with a plan
# change (plan_change = True) are shaded orange. A clear info /
# legend block is rendered at the top of every tab.
#
# Output (one file per method):
#   {gt_root}/summaries_nc/qerr_details_{method}_sf{sf}_{res}_{run_type}_nc.xlsx
#
# Usage:
#   python scripts/excel_qerr_desc_by_method_nc.py                       # default folder
#   python scripts/excel_qerr_desc_by_method_nc.py gt_results_sf1_10x10_s1q0
# =========================================================

import re
import sys
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


DEFAULT_GT_ROOT = "gt_results_sf1_10x10_s1q0"

_GT_RE = re.compile(
    r"^gt_results_sf(?P<sf>\d+)_(?P<res>\d+x\d+)_s(?P<sys>\d+)q(?P<q>\d+)(?P<ma>_ma)?$",
    re.IGNORECASE,
)

METHOD_DESC = {
    "m0": "data-space uniform resolution",
    "m1": "selectivity-space uniform resolution (percentile)",
    "m2": "selectivity-space exponential resolution (geometric / Picasso)",
}
METHOD_KIND = {
    "m0": "Uniform (data space)",
    "m1": "Uniform (selectivity space)",
    "m2": "Exponential (selectivity space)",
}
METHOD_COLORS = {"m0": "2E75B6", "m1": "548235", "m2": "C55A11"}

# palette
C_TITLE = "1F3864"     # dark navy
C_LBL   = "D9E1F2"     # light blue label
C_HDR   = "4472C4"     # table header blue
C_BAND  = "EDF2FB"     # row banding
C_PLAN  = "F8CBAD"     # plan-change row highlight (orange)
WHITE   = "FFFFFF"
GREY    = "404040"

_THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _fill(color):
    return PatternFill("solid", fgColor=color)


# =========================================================
# folder metadata
# =========================================================

def _parse_root(name):
    m = _GT_RE.match(name)
    if not m:
        return {"sf": "?", "res": "?", "sys": 1, "q": 0, "ma": False, "run_type": name}
    sysw = int(m.group("sys"))
    qw = int(m.group("q"))
    ma = bool(m.group("ma"))
    return {
        "sf": m.group("sf"),
        "res": m.group("res"),
        "sys": sysw,
        "q": qw,
        "ma": ma,
        "run_type": f"s{sysw}q{qw}" + ("_ma" if ma else ""),
    }


def _resolution_note(res):
    try:
        parts = [int(p) for p in res.lower().split("x")]
        pts = "×".join(str(p + 1) for p in parts)
        return f"{res}   (gaps between {pts} sampled grid points per axis)"
    except Exception:
        return res


def _meta(info, queries, method, query, nrows):
    sysp = (
        "OFF — 1 system worker"
        if info["sys"] == 1
        else f"ON — {info['sys']} system workers"
    )
    qp = (
        "OFF — max_parallel_workers_per_gather = 0"
        if info["q"] == 0
        else f"ON — max_parallel_workers_per_gather = {info['q']}"
    )
    ma = "ON" if info["ma"] else "OFF — initial min point at 0 selectivity"

    return [
        ("Data scope", "Non-zero cardinality only (_nc) — instances with count_rows = 0 OR neighbor_count_rows = 0 are removed"),
        ("Folder", f"gt_results_sf{info['sf']}_{info['res']}_{info['run_type']}"),
        ("Scale factor (SF)", info["sf"]),
        ("Resolution", _resolution_note(info["res"])),
        ("System parallelism", sysp),
        ("Query parallelism", qp),
        ("Min-adjustment (_ma)", ma),
        ("Total queries", f"{len(queries)}   ({', '.join(q.upper() for q in queries)})"),
        ("Method", f"{method.upper()} — {METHOD_DESC.get(method, method)}   [{METHOD_KIND.get(method, '')}]"),
        ("This tab", f"Query {query.upper()} — {nrows} rows"),
        ("Highlight", "Rows shaded orange = plan change (plan_change = True)"),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]


# =========================================================
# value formatting
# =========================================================

def _coerce(value):
    if value is None:
        return None, None
    if isinstance(value, float) and np.isnan(value):
        return None, None
    try:
        if pd.isna(value):
            return None, None
    except (TypeError, ValueError):
        pass

    if isinstance(value, (bool, np.bool_)):
        return bool(value), None
    if isinstance(value, (int, np.integer)):
        return int(value), "#,##0"
    if isinstance(value, (float, np.floating)):
        f = float(value)
        if f == int(f) and abs(f) < 1e15:
            return int(f), "#,##0"
        return f, "#,##0.000000"
    return str(value), None


# =========================================================
# low-level writers
# =========================================================

def _title(ws, row, ncols, text, color=C_TITLE, size=15, height=26):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color=WHITE, size=size)
    c.fill = _fill(color)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = height
    return row + 1


def _meta_block(ws, row, ncols, items):
    for label, value in items:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(bold=True, color=GREY, size=10)
        lc.fill = _fill(C_LBL)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lc.border = BORDER

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=ncols)
        vc = ws.cell(row=row, column=2, value=str(value))
        vc.font = Font(color="222222", size=10)
        vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        vc.border = BORDER
        ws.row_dimensions[row].height = 16
        row += 1
    return row


def _section_header(ws, row, ncols, text, color):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color=WHITE, size=12)
    c.fill = _fill(color)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20
    return row + 1


def _table_header(ws, row, cols, color=C_HDR):
    for j, name in enumerate(cols, start=1):
        c = ws.cell(row=row, column=j, value=name)
        c.font = Font(bold=True, color=WHITE, size=9)
        c.fill = _fill(color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 28
    return row + 1


def _is_plan_change(v):
    return str(v).strip().lower() in {"true", "t", "1", "yes", "structural"}


def _table_rows(ws, row, df, cols):
    band = _fill(C_BAND)
    plan = _fill(C_PLAN)
    for i, (_, r) in enumerate(df.iterrows()):
        changed = _is_plan_change(r.get("plan_change"))
        rowfill = plan if changed else (band if i % 2 == 1 else None)
        for j, name in enumerate(cols, start=1):
            val, nf = _coerce(r.get(name))
            c = ws.cell(row=row, column=j, value=val)
            c.border = BORDER
            c.font = Font(size=9, bold=changed)
            if nf:
                c.number_format = nf
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
            if rowfill is not None:
                c.fill = rowfill
        row += 1
    return row


def _autosize(ws, cols, df):
    for j, name in enumerate(cols, start=1):
        width = len(str(name))
        if name in df.columns:
            sample = df[name].astype(str).head(200)
            if len(sample):
                width = max(width, int(sample.map(len).max()))
        ws.column_dimensions[get_column_letter(j)].width = min(max(width + 2, 8), 44)


# =========================================================
# discovery
# =========================================================

def _discover(gt_root):
    """Return {query: {method: csv_path}} for qerr_desc_nc.csv files."""
    out = {}
    for qdir in sorted(gt_root.iterdir()):
        if not qdir.is_dir() or qdir.name.startswith("summaries"):
            continue
        methods = {}
        for mdir in sorted(qdir.iterdir()):
            if not mdir.is_dir():
                continue
            p = mdir / "qerr_sorted" / "qerr_desc_nc.csv"
            if p.exists():
                methods[mdir.name] = p
        if methods:
            out[qdir.name] = methods
    return out


# =========================================================
# main
# =========================================================

def run(gt_root):
    gt_root = Path(gt_root).resolve()
    info = _parse_root(gt_root.name)

    print()
    print("=" * 70)
    print("EXCEL — qerr_desc_nc (one workbook PER METHOD, per-query tabs)")
    print("=" * 70)
    print(f"Folder : {gt_root.name}")

    discovered = _discover(gt_root)
    if not discovered:
        print(f"No qerr_desc_nc.csv found under: {gt_root}")
        return

    queries = sorted(discovered)
    methods = sorted({m for mp in discovered.values() for m in mp})
    print(f"Queries: {queries}")
    print(f"Methods: {methods}")

    out_dir = gt_root / "summaries_nc"
    out_dir.mkdir(parents=True, exist_ok=True)

    for method in methods:
        wb = Workbook()
        wb.remove(wb.active)

        method_queries = [q for q in queries if method in discovered[q]]

        for query in method_queries:
            df = pd.read_csv(discovered[query][method])
            cols = list(df.columns)
            ncols = max(len(cols), 2)

            ws = wb.create_sheet(title=query.upper()[:31])
            ws.sheet_view.showGridLines = False

            row = 1
            row = _title(
                ws, row, ncols,
                f"{query.upper()}  ·  METHOD {method.upper()} — Q-error instances  ·  "
                f"NON-ZERO CARDINALITY (_nc)",
                color=METHOD_COLORS.get(method, C_TITLE),
            )
            row += 1
            row = _meta_block(ws, row, ncols, _meta(info, queries, method, query, len(df)))
            row += 1

            row = _section_header(
                ws, row, ncols,
                f"METHOD {method.upper()}  —  {METHOD_DESC.get(method, method)}   "
                f"[{METHOD_KIND.get(method, '')}]   ·   {len(df)} rows",
                METHOD_COLORS.get(method, "595959"),
            )
            row = _table_header(ws, row, cols)
            row = _table_rows(ws, row, df, cols)

            _autosize(ws, cols, df)
            # title block scrollable (no freeze)
            ws.freeze_panes = None

        out_path = out_dir / (
            f"qerr_details_{method}_sf{info['sf']}_{info['res']}_{info['run_type']}_nc.xlsx"
        )
        wb.save(out_path)
        print(f"  Saved {method.upper()} ({len(method_queries)} tabs): {out_path.name}")

    print()
    print(f"Output dir: {out_dir}")
    print("=" * 70)


if __name__ == "__main__":
    base = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_GT_ROOT
    run(base)
