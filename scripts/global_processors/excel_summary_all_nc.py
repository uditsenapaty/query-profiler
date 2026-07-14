# =========================================================
# scripts/excel_summary_all_nc.py
#
# Builds ONE beautiful .xlsx for a single gt_results_* folder,
# replacing the 3 per-method summary_all CSVs in summaries_nc/:
#   - 3 tabs, one per method (m0, m1, m2)
#   - each tab = that method's cross-query summary table
#     (Metric x query), with section banners
#
# Only _nc (non-zero cardinality) data is used. A clear info /
# legend block is rendered at the top of every tab stating:
#   total queries, system/query parallelism on/off, resolution
#   (with the 11x11-points note), the method's sampling kind
#   (uniform / exponential), min-adjustment, and the nc scope.
#
# Output:
#   {gt_root}/summaries_nc/summary_all_nc_{run_type}.xlsx
#
# Usage:
#   python scripts/excel_summary_all_nc.py                       # default folder
#   python scripts/excel_summary_all_nc.py gt_results_sf1_10x10_s1q0
# =========================================================

import re
import sys
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


DEFAULT_GT_ROOT = "gt_results_sf1_10x10_s1q0"

_GT_RE = re.compile(
    r"^gt_results_sf(?P<sf>\d+)_(?P<res>\d+x\d+)_s(?P<sys>\d+)q(?P<q>\d+)(?P<ma>_ma)?$",
    re.IGNORECASE,
)
_FILE_RE = re.compile(r"^summary_all_(?P<method>m\d+)_.*_nc\.csv$", re.IGNORECASE)

METHOD_DESC = {
    "m0": "data-space uniform resolution",
    "m1": "selectivity-space uniform resolution (percentile)",
    "m2": "selectivity-space exponential resolution (geometric / Picasso)",
}
METHOD_KIND = {
    "m0": "Uniform (data space)",
    "m1": "Uniform (selectivity space)",
    "m2": "Exponential (selectivity space)",
}
METHOD_COLORS = {"m0": "2E75B6", "m1": "548235", "m2": "C55A11"}

# palette
C_TITLE   = "1F3864"   # dark navy
C_LBL     = "D9E1F2"   # light blue label
C_HDR     = "4472C4"   # table header blue
C_SECTION = "8EAADB"   # section banner
C_BAND    = "EDF2FB"   # row banding
WHITE     = "FFFFFF"
GREY      = "404040"

_THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _fill(color):
    return PatternFill("solid", fgColor=color)


# =========================================================
# folder metadata
# =========================================================

def _parse_root(name):
    m = _GT_RE.match(name)
    if not m:
        return {"sf": "?", "res": "?", "sys": 1, "q": 0, "ma": False, "run_type": name}
    sysw = int(m.group("sys"))
    qw = int(m.group("q"))
    ma = bool(m.group("ma"))
    return {
        "sf": m.group("sf"),
        "res": m.group("res"),
        "sys": sysw,
        "q": qw,
        "ma": ma,
        "run_type": f"s{sysw}q{qw}" + ("_ma" if ma else ""),
    }


def _resolution_note(res):
    try:
        parts = [int(p) for p in res.lower().split("x")]
        pts = "×".join(str(p + 1) for p in parts)
        return f"{res}   (gaps between {pts} sampled grid points per axis)"
    except Exception:
        return res


def _global_meta(info, queries, method):
    sysp = (
        "OFF — 1 system worker"
        if info["sys"] == 1
        else f"ON — {info['sys']} system workers"
    )
    qp = (
        "OFF — max_parallel_workers_per_gather = 0"
        if info["q"] == 0
        else f"ON — max_parallel_workers_per_gather = {info['q']}"
    )
    ma = "ON" if info["ma"] else "OFF — initial min point at 0 selectivity"

    return [
        ("Data scope", "Non-zero cardinality only (_nc) — instances with count_rows = 0 OR neighbor_count_rows = 0 are removed"),
        ("Folder", f"gt_results_sf{info['sf']}_{info['res']}_{info['run_type']}"),
        ("Scale factor (SF)", info["sf"]),
        ("Resolution", _resolution_note(info["res"])),
        ("System parallelism", sysp),
        ("Query parallelism", qp),
        ("Min-adjustment (_ma)", ma),
        ("Total queries", f"{len(queries)}   ({', '.join(q.upper() for q in queries)})"),
        ("Method", f"{method.upper()} — {METHOD_DESC.get(method, method)}   [{METHOD_KIND.get(method, '')}]"),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]


# =========================================================
# value formatting
# =========================================================

def _coerce_str(s):
    """Numeric strings -> numbers (right aligned); everything else stays text."""
    s = "" if s is None else str(s).strip()
    if s == "":
        return None, None
    try:
        f = float(s)
    except ValueError:
        return s, None
    if "." not in s and "e" not in s.lower() and abs(f) < 1e15:
        return int(f), "#,##0"
    return f, "#,##0.000000"


# =========================================================
# low-level writers
# =========================================================

def _title(ws, row, ncols, text, color=C_TITLE, size=15, height=26):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color=WHITE, size=size)
    c.fill = _fill(color)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = height
    return row + 1


def _meta_block(ws, row, ncols, items):
    for label, value in items:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(bold=True, color=GREY, size=10)
        lc.fill = _fill(C_LBL)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lc.border = BORDER

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=ncols)
        vc = ws.cell(row=row, column=2, value=str(value))
        vc.font = Font(color="222222", size=10)
        vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        vc.border = BORDER
        ws.row_dimensions[row].height = 16
        row += 1
    return row


def _table_header(ws, row, headers, color=C_HDR):
    for j, name in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=name)
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = _fill(color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 22
    return row + 1


def _section_banner(ws, row, ncols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color=C_TITLE, size=11)
    c.fill = _fill(C_SECTION)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = BORDER
    ws.row_dimensions[row].height = 19
    return row + 1


# =========================================================
# discovery
# =========================================================

def _discover(summaries_dir, run_type):
    """Return {method: csv_path} for summary_all_*_nc.csv."""
    out = {}
    if not summaries_dir.is_dir():
        return out
    for p in sorted(summaries_dir.glob("summary_all_*_nc.csv")):
        m = _FILE_RE.match(p.name)
        if not m:
            continue
        # only this folder's run_type
        if f"_{run_type}_nc.csv" not in p.name:
            continue
        out[m.group("method").lower()] = p
    return out


# =========================================================
# main
# =========================================================

def run(gt_root):
    gt_root = Path(gt_root).resolve()
    info = _parse_root(gt_root.name)
    summaries_dir = gt_root / "summaries_nc"

    print()
    print("=" * 70)
    print("EXCEL — summary_all_nc (per-method tabs)")
    print("=" * 70)
    print(f"Folder : {gt_root.name}")

    found = _discover(summaries_dir, info["run_type"])
    if not found:
        print(f"No summary_all_*_{info['run_type']}_nc.csv found under: {summaries_dir}")
        return

    methods = sorted(found)
    print(f"Methods: {methods}")

    wb = Workbook()
    wb.remove(wb.active)

    queries_seen = []

    for method in methods:
        df = pd.read_csv(found[method], dtype=str).fillna("")

        query_cols = [c for c in df.columns if c not in ("Section", "Metric")]
        if not queries_seen:
            queries_seen = [q.lower() for q in query_cols]

        headers = ["Metric"] + [q.upper() for q in query_cols]
        ncols = len(headers)

        ws = wb.create_sheet(title=method.upper()[:31])
        ws.sheet_view.showGridLines = False

        row = 1
        row = _title(
            ws, row, ncols,
            f"METHOD {method.upper()} — {METHOD_DESC.get(method, method)}  ·  "
            f"NON-ZERO CARDINALITY (_nc)",
            color=METHOD_COLORS.get(method, C_TITLE),
        )
        row += 1
        row = _meta_block(
            ws, row, ncols,
            _global_meta(info, [q.lower() for q in query_cols], method),
        )
        row += 1

        header_row = row
        row = _table_header(ws, row, headers)

        band = _fill(C_BAND)
        data_i = 0
        for _, r in df.iterrows():
            section = str(r["Section"]).strip()
            metric = str(r["Metric"]).strip()

            # section banner row
            if metric == "" and section != "":
                row = _section_banner(ws, row, ncols, section)
                data_i = 0
                continue

            if metric == "":
                continue

            # data row: Metric + per-query values
            mc = ws.cell(row=row, column=1, value=metric)
            mc.font = Font(bold=True, size=9, color="222222")
            mc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            mc.border = BORDER
            if data_i % 2 == 1:
                mc.fill = band

            for j, qc in enumerate(query_cols, start=2):
                val, nf = _coerce_str(r[qc])
                c = ws.cell(row=row, column=j, value=val)
                c.border = BORDER
                c.font = Font(size=9)
                if nf:
                    c.number_format = nf
                    c.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center")
                if data_i % 2 == 1:
                    c.fill = band

            data_i += 1
            row += 1

        # widths
        ws.column_dimensions["A"].width = 26
        for j in range(2, ncols + 1):
            ws.column_dimensions[get_column_letter(j)].width = 16

        # No freeze: keep the whole sheet (incl. title block) scrollable.
        ws.freeze_panes = None
        print(f"  + tab {method.upper()}  ({len(query_cols)} queries)")

    out_path = summaries_dir / (
        f"summary_per_method_sf{info['sf']}_{info['res']}_{info['run_type']}_nc.xlsx"
    )
    wb.save(out_path)

    print()
    print(f"Saved: {out_path}")
    print("=" * 70)


if __name__ == "__main__":
    base = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_GT_ROOT
    run(base)
