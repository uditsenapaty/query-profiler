# =========================================================
# scripts/global_processors/summary_global.py
#
# Collects the per-method summary.csv files of a single query into
# ONE workbook, one tab per method — written inside the query folder.
#
# Called by build_gt.py as a global processor (once per query):
#   run(config_gt.RESULTS_DIR.parent)          # a single query dir
#
# Input layout:
#   {query_dir}/m0/summary.csv
#   {query_dir}/m1/summary.csv
#   {query_dir}/m2/summary.csv
#
# Output (one file, tabs = methods, no CSVs):
#   {query_dir}/summary_all_methods.xlsx
# =========================================================

from pathlib import Path
import re

import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


SUMMARY_NAME = "summary.csv"                 # per-method input
OUT_NAME     = "summary_all_methods.xlsx"    # per-query output (tabs = methods)


def _natural_key(name: str):
    """Sort m0 < m1 < m2 (and qt5 < qt10)."""
    m = re.search(r"(\d+)", name)
    return (re.sub(r"\d+", "", name), int(m.group(1)) if m else -1, name)


def _discover_methods(query_dir: Path) -> dict:
    """{method_name: method_dir} for every subdir that has a summary."""
    found = {}
    for d in sorted(query_dir.iterdir(), key=lambda p: _natural_key(p.name)):
        if d.is_dir() and (d / SUMMARY_NAME).exists():
            found[d.name] = d
    return found


def _dump_sheet(ws, df: pd.DataFrame):
    """Write a method's summary (Section/Metric/Value) into a sheet, as-is."""
    # header
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=j, value=str(col))
        c.font = Font(bold=True)

    # rows (blank cells left empty)
    for i, (_, r) in enumerate(df.iterrows(), start=2):
        for j, col in enumerate(df.columns, start=1):
            v = r[col]
            ws.cell(row=i, column=j, value=(v if v != "" else None))

    ncols = len(df.columns)
    ws.column_dimensions["A"].width = 34
    if ncols >= 2:
        ws.column_dimensions["B"].width = 42
    for j in range(3, ncols + 1):
        ws.column_dimensions[get_column_letter(j)].width = 22
    ws.freeze_panes = "A2"


def run(results_dir):

    query_dir = Path(results_dir)          # a single query dir

    print()
    print("=" * 60)
    print("SUMMARY GLOBAL — ONE WORKBOOK, ONE TAB PER METHOD")
    print("=" * 60)
    print(f"Query dir : {query_dir}")

    method_dirs = _discover_methods(query_dir)

    if not method_dirs:
        print(f"No {SUMMARY_NAME} found under:\n  {query_dir}")
        return

    print(f"Methods   : {', '.join(method_dirs)}")

    wb = Workbook()
    wb.remove(wb.active)

    for method in sorted(method_dirs, key=_natural_key):
        try:
            df = pd.read_csv(method_dirs[method] / SUMMARY_NAME, dtype=str).fillna("")
        except Exception as e:
            print(f"  [WARN] skipping {method}: {e}")
            continue

        ws = wb.create_sheet(title=method[:31])
        _dump_sheet(ws, df)
        print(f"  + tab {method}  ({len(df)} rows)")

    if not wb.sheetnames:
        print("No valid summaries loaded.")
        return

    out_path = query_dir / OUT_NAME
    wb.save(out_path)

    print(f"Saved     : {out_path}")
    print("DONE")


if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "gt_results_sf1_10x10_s1q0/qt8"
    run(path)
